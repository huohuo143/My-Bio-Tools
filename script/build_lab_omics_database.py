#!/usr/bin/env python3
"""Build the Wu Lab analysed-omics registry and query database.

Only already analysed tables are read. Source directories are never modified;
all derived files are written below the requested output directory.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import asdict, dataclass
import gzip
import hashlib
import json
import math
import os
from pathlib import Path
import re
import shutil
import sqlite3
import subprocess
import sys
from typing import Iterable, Iterator, Sequence

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_OUTPUT = Path("/Volumes/FAFU/analysis_results/wulab_omics_app_v1")
REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MAPPING = (
    REPO_ROOT
    / "app_source/data/Rice_Genome_Annotation_Project/RAP-MSU_2025-03-19.txt.gz"
)

MSU_PATTERN = re.compile(r"LOC_Os\d{2}g\d{5}(?:\.\d+)?", re.IGNORECASE)
RAP_GENE_PATTERN = re.compile(r"Os\d{2}g\d{7}", re.IGNORECASE)
RAP_MODEL_PATTERN = re.compile(r"Os\d{2}t\d{7}(?:-\d+)?", re.IGNORECASE)


@dataclass(frozen=True)
class MappingRecord:
    original_id: str
    msu_locus: str = ""
    msu_model: str = ""
    rap_gene: str = ""
    rap_model: str = ""
    status: str = "unmapped"
    note: str = ""


def clean_text(value: object, limit: int = 2_000) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.casefold() in {"nan", "none", "na", "--"}:
        return ""
    return text[:limit]


def as_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def json_values(values: Sequence[object]) -> str:
    cleaned = [as_float(value) for value in values]
    return json.dumps(cleaned, ensure_ascii=False, separators=(",", ":"))


def add_direction_qc(
    builder: "OmicsDatabaseBuilder",
    dataset_id: str,
    label: str,
    source_effect: Sequence[object],
    derived_effect: Sequence[object],
) -> None:
    pairs = [
        (left, right)
        for left, right in zip((as_float(value) for value in source_effect), (as_float(value) for value in derived_effect))
        if left is not None and right is not None and left != 0 and right != 0
    ]
    if not pairs:
        builder.add_qc(dataset_id, f"{label}_direction_concordance", text="no comparable rows", status="review")
        return
    left = pd.Series([item[0] for item in pairs], dtype=float)
    right = pd.Series([item[1] for item in pairs], dtype=float)
    concordance = float(((left > 0) == (right > 0)).mean())
    correlation = float(left.corr(right)) if len(pairs) > 2 else None
    status = "pass" if concordance >= 0.95 else "review"
    builder.add_qc(dataset_id, f"{label}_direction_concordance", concordance, status=status, notes=f"n={len(pairs)}; derived from treatment/control source abundance")
    builder.add_qc(dataset_id, f"{label}_effect_pearson", correlation, status="pass" if correlation is not None and correlation >= 0.9 else "review", notes=f"n={len(pairs)}")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


class IdentifierResolver:
    """Resolve RAP/MSU identifiers while preserving every one-to-many mapping."""

    def __init__(self, mapping_path: Path) -> None:
        self.mapping_path = mapping_path
        self.rap_to_models: dict[str, list[str]] = {}
        self.model_to_rap: dict[str, list[str]] = {}
        self.locus_to_models: dict[str, list[str]] = {}
        self.locus_to_rap: dict[str, list[str]] = {}
        with gzip.open(mapping_path, "rt", encoding="utf-8") as handle:
            for line in handle:
                rap, raw_models = line.rstrip("\n").split("\t", 1)
                rap = rap.strip()
                models = [] if raw_models == "None" else [item.strip() for item in raw_models.split(",") if item.strip()]
                self.rap_to_models[rap.casefold()] = models
                for model in models:
                    locus = model.split(".", 1)[0]
                    self.model_to_rap.setdefault(model.casefold(), []).append(rap)
                    self.locus_to_models.setdefault(locus.casefold(), []).append(model)
                    self.locus_to_rap.setdefault(locus.casefold(), []).append(rap)
        for index in (self.model_to_rap, self.locus_to_models, self.locus_to_rap):
            for key, values in index.items():
                index[key] = list(dict.fromkeys(values))

    def resolve(self, original_id: object, fallbacks: Sequence[object] = ()) -> list[MappingRecord]:
        original = clean_text(original_id, 500)
        candidates = [original, *(clean_text(value, 500) for value in fallbacks)]
        msu_hits: list[str] = []
        rap_gene_hits: list[str] = []
        rap_model_hits: list[str] = []
        for candidate in candidates:
            msu_hits.extend(match.group(0) for match in MSU_PATTERN.finditer(candidate))
            rap_gene_hits.extend(match.group(0) for match in RAP_GENE_PATTERN.finditer(candidate))
            rap_model_hits.extend(match.group(0) for match in RAP_MODEL_PATTERN.finditer(candidate))
        msu_hits = list(dict.fromkeys(msu_hits))
        rap_gene_hits = list(dict.fromkeys(rap_gene_hits))
        rap_model_hits = list(dict.fromkeys(rap_model_hits))
        for rap_model in rap_model_hits:
            rap_gene_hits.append(rap_model.replace("t", "g", 1).split("-", 1)[0])
        rap_gene_hits = list(dict.fromkeys(rap_gene_hits))

        model_candidates: list[str] = []
        for hit in msu_hits:
            if "." in hit:
                model_candidates.append(hit)
            else:
                model_candidates.extend(self.locus_to_models.get(hit.casefold(), []))
                if not self.locus_to_models.get(hit.casefold()):
                    model_candidates.append("")
        for rap in rap_gene_hits:
            model_candidates.extend(self.rap_to_models.get(rap.casefold(), []))
        model_candidates = list(dict.fromkeys(model_candidates))

        records: list[MappingRecord] = []
        if model_candidates:
            loci = {
                (model.split(".", 1)[0] if model else hit.split(".", 1)[0])
                for model in model_candidates
                for hit in (msu_hits[:1] or [""])
                if model or hit
            }
            locus_count = len(loci)
            status = "multi_locus" if locus_count > 1 else "multi_model" if len(model_candidates) > 1 else "mapped"
            for model in model_candidates:
                locus = model.split(".", 1)[0] if model else (msu_hits[0].split(".", 1)[0] if msu_hits else "")
                raps = self.model_to_rap.get(model.casefold(), []) if model else self.locus_to_rap.get(locus.casefold(), [])
                if not raps:
                    raps = rap_gene_hits or [""]
                for rap in raps:
                    records.append(
                        MappingRecord(
                            original_id=original,
                            msu_locus=locus,
                            msu_model=model,
                            rap_gene=rap,
                            rap_model=rap_model_hits[0] if rap_model_hits else "",
                            status=status,
                            note="RAP-MSU_2025-03-19 one-to-many preserved",
                        )
                    )
        elif msu_hits:
            for hit in msu_hits:
                locus = hit.split(".", 1)[0]
                records.append(
                    MappingRecord(
                        original_id=original,
                        msu_locus=locus,
                        msu_model=hit if "." in hit else "",
                        rap_gene=(self.locus_to_rap.get(locus.casefold(), [""]) or [""])[0],
                        rap_model=rap_model_hits[0] if rap_model_hits else "",
                        status="mapped",
                        note="MSU locus retained; no model-level mapping in reference" if "." not in hit else "direct MSU model",
                    )
                )
        if not records:
            records = [MappingRecord(original_id=original, rap_gene=rap_gene_hits[0] if rap_gene_hits else "", rap_model=rap_model_hits[0] if rap_model_hits else "")]
        return list(dict.fromkeys(records))


SCHEMA = """
PRAGMA foreign_keys=ON;
CREATE TABLE metadata (key TEXT PRIMARY KEY, value TEXT NOT NULL);
CREATE TABLE datasets (
  dataset_id TEXT PRIMARY KEY, display_name TEXT NOT NULL, category TEXT NOT NULL,
  assay TEXT NOT NULL, host_background TEXT, treatment TEXT, control_group TEXT,
  organism TEXT NOT NULL DEFAULT 'Oryza sativa', inclusion_status TEXT NOT NULL,
  inclusion_reason TEXT, source_root TEXT, replicate_note TEXT, descriptive INTEGER NOT NULL DEFAULT 0,
  historical INTEGER NOT NULL DEFAULT 0, id_namespace TEXT, abundance_unit TEXT,
  time_order TEXT, notes TEXT
);
CREATE TABLE samples (
  dataset_id TEXT NOT NULL, sample_order INTEGER NOT NULL, sample_id TEXT NOT NULL,
  original_sample_code TEXT NOT NULL, group_id TEXT, condition_role TEXT,
  treatment TEXT, time_label TEXT, time_order INTEGER, replicate INTEGER,
  unit TEXT, PRIMARY KEY(dataset_id, sample_id)
);
CREATE TABLE comparisons (
  comparison_id TEXT PRIMARY KEY, dataset_id TEXT NOT NULL, display_name TEXT NOT NULL,
  treatment_group TEXT NOT NULL, control_group TEXT NOT NULL, treatment TEXT,
  time_label TEXT, time_order INTEGER, n_treatment INTEGER, n_control INTEGER,
  direction TEXT NOT NULL, effect_metric TEXT NOT NULL, p_metric TEXT,
  descriptive INTEGER NOT NULL DEFAULT 0, source_file TEXT, source_sheet TEXT, notes TEXT
);
CREATE TABLE files (
  file_id INTEGER PRIMARY KEY AUTOINCREMENT, dataset_id TEXT, path TEXT NOT NULL UNIQUE,
  role TEXT NOT NULL, exists_flag INTEGER NOT NULL, size_bytes INTEGER, sha256 TEXT,
  modified_at TEXT, source_format TEXT, duplicate_of TEXT, notes TEXT
);
CREATE TABLE gene_mappings (
  mapping_id INTEGER PRIMARY KEY AUTOINCREMENT, dataset_id TEXT NOT NULL,
  original_id TEXT NOT NULL, msu_locus TEXT, msu_model TEXT, rap_gene TEXT, rap_model TEXT,
  mapping_status TEXT NOT NULL, mapping_note TEXT,
  UNIQUE(dataset_id, original_id, msu_locus, msu_model, rap_gene, rap_model)
);
CREATE TABLE differential_results (
  result_id INTEGER PRIMARY KEY AUTOINCREMENT, dataset_id TEXT NOT NULL,
  comparison_id TEXT NOT NULL, original_id TEXT NOT NULL, msu_locus TEXT NOT NULL,
  msu_model TEXT, rap_gene TEXT, rap_model TEXT, feature_type TEXT NOT NULL,
  log2fc REAL, ratio REAL, pvalue REAL, padj REAL, regulated TEXT,
  annotation TEXT, protein_accession TEXT, site_position INTEGER, site_residue TEXT,
  modified_sequence TEXT, localization_probability REAL, descriptive INTEGER NOT NULL DEFAULT 0,
  source_file TEXT NOT NULL, source_sheet TEXT, source_row INTEGER
);
CREATE TABLE abundance_profiles (
  profile_id INTEGER PRIMARY KEY AUTOINCREMENT, dataset_id TEXT NOT NULL,
  original_id TEXT NOT NULL, msu_locus TEXT NOT NULL, msu_model TEXT,
  rap_gene TEXT, rap_model TEXT, feature_type TEXT NOT NULL,
  site_position INTEGER, site_residue TEXT, modified_sequence TEXT,
  unit TEXT NOT NULL, values_json TEXT NOT NULL, source_file TEXT NOT NULL,
  source_sheet TEXT, source_row INTEGER
);
CREATE TABLE qc_metrics (
  metric_id INTEGER PRIMARY KEY AUTOINCREMENT, dataset_id TEXT,
  metric TEXT NOT NULL, value REAL, text_value TEXT, status TEXT, notes TEXT
);
CREATE INDEX idx_diff_msu ON differential_results(msu_locus);
CREATE INDEX idx_diff_dataset_msu ON differential_results(dataset_id, msu_locus);
CREATE INDEX idx_profile_msu ON abundance_profiles(msu_locus);
CREATE INDEX idx_profile_dataset_msu ON abundance_profiles(dataset_id, msu_locus);
CREATE INDEX idx_mapping_original ON gene_mappings(original_id);
CREATE INDEX idx_mapping_msu ON gene_mappings(msu_locus);
"""


class OmicsDatabaseBuilder:
    def __init__(self, db_path: Path, resolver: IdentifierResolver, row_limit: int | None = None) -> None:
        self.db_path = db_path
        self.resolver = resolver
        self.row_limit = row_limit
        self.connection = sqlite3.connect(db_path)
        self.connection.executescript(SCHEMA)
        self.mapping_cache: dict[tuple[str, str, tuple[str, ...]], list[MappingRecord]] = {}
        self.file_cache: set[str] = set()
        self.connection.executemany(
            "INSERT INTO metadata(key,value) VALUES (?,?)",
            [
                ("schema_version", "1"),
                ("app_target", "My Bio Tools v1.9.1"),
                ("primary_gene_key", "MSU locus without model suffix"),
                ("raw_data_reanalysed", "false"),
                ("mapping_reference", str(resolver.mapping_path)),
            ],
        )

    def close(self) -> None:
        self.connection.commit()
        self.connection.execute("ANALYZE")
        self.connection.execute("PRAGMA optimize")
        self.connection.close()

    def add_dataset(self, **row: object) -> None:
        columns = [
            "dataset_id", "display_name", "category", "assay", "host_background", "treatment",
            "control_group", "organism", "inclusion_status", "inclusion_reason", "source_root",
            "replicate_note", "descriptive", "historical", "id_namespace", "abundance_unit",
            "time_order", "notes",
        ]
        values = [row.get(column, "") for column in columns]
        values[8] = row.get("inclusion_status", "included")
        values[3] = row.get("assay", "")
        values[2] = row.get("category", "")
        values[1] = row.get("display_name", row.get("dataset_id", ""))
        values[7] = row.get("organism", "Oryza sativa")
        values[12] = int(bool(row.get("descriptive", 0)))
        values[13] = int(bool(row.get("historical", 0)))
        self.connection.execute(
            f"INSERT OR REPLACE INTO datasets({','.join(columns)}) VALUES ({','.join('?' for _ in columns)})",
            values,
        )

    def add_samples(self, dataset_id: str, rows: Sequence[dict[str, object]]) -> None:
        self.connection.executemany(
            """INSERT OR REPLACE INTO samples(
              dataset_id,sample_order,sample_id,original_sample_code,group_id,condition_role,
              treatment,time_label,time_order,replicate,unit
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            [
                (
                    dataset_id, index, row["sample_id"], row.get("original_sample_code", row["sample_id"]),
                    row.get("group_id", ""), row.get("condition_role", ""), row.get("treatment", ""),
                    row.get("time_label", ""), row.get("time_order"), row.get("replicate"), row.get("unit", ""),
                )
                for index, row in enumerate(rows)
            ],
        )

    def add_comparison(self, **row: object) -> None:
        columns = [
            "comparison_id", "dataset_id", "display_name", "treatment_group", "control_group",
            "treatment", "time_label", "time_order", "n_treatment", "n_control", "direction",
            "effect_metric", "p_metric", "descriptive", "source_file", "source_sheet", "notes",
        ]
        values = [row.get(column, "") for column in columns]
        values[8] = row.get("n_treatment")
        values[9] = row.get("n_control")
        values[13] = int(bool(row.get("descriptive", 0)))
        self.connection.execute(
            f"INSERT OR REPLACE INTO comparisons({','.join(columns)}) VALUES ({','.join('?' for _ in columns)})",
            values,
        )

    def add_file(self, dataset_id: str, path: Path, role: str, *, duplicate_of: str = "", notes: str = "") -> None:
        key = str(path)
        if key in self.file_cache:
            return
        self.file_cache.add(key)
        exists = path.is_file()
        stat = path.stat() if exists else None
        digest = sha256_file(path) if exists else ""
        self.connection.execute(
            """INSERT INTO files(dataset_id,path,role,exists_flag,size_bytes,sha256,modified_at,source_format,duplicate_of,notes)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                dataset_id, key, role, int(exists), stat.st_size if stat else None, digest,
                str(pd.Timestamp(stat.st_mtime, unit="s", tz="UTC")) if stat else "",
                path.suffix.lower().lstrip("."), duplicate_of, notes,
            ),
        )

    def add_qc(self, dataset_id: str, metric: str, value: float | None = None, *, text: str = "", status: str = "info", notes: str = "") -> None:
        self.connection.execute(
            "INSERT INTO qc_metrics(dataset_id,metric,value,text_value,status,notes) VALUES (?,?,?,?,?,?)",
            (dataset_id, metric, value, text, status, notes),
        )

    def mappings(self, dataset_id: str, original_id: object, fallbacks: Sequence[object] = ()) -> list[MappingRecord]:
        original = clean_text(original_id, 500)
        fallback_text = tuple(clean_text(value, 500) for value in fallbacks if clean_text(value, 500))
        key = (dataset_id, original, fallback_text)
        if key in self.mapping_cache:
            return self.mapping_cache[key]
        records = self.resolver.resolve(original, fallback_text)
        self.connection.executemany(
            """INSERT OR IGNORE INTO gene_mappings(
              dataset_id,original_id,msu_locus,msu_model,rap_gene,rap_model,mapping_status,mapping_note
            ) VALUES (?,?,?,?,?,?,?,?)""",
            [
                (dataset_id, item.original_id, item.msu_locus, item.msu_model, item.rap_gene, item.rap_model, item.status, item.note)
                for item in records
            ],
        )
        self.mapping_cache[key] = records
        return records

    def add_profile(
        self,
        dataset_id: str,
        original_id: object,
        values: Sequence[object],
        *,
        feature_type: str,
        unit: str,
        source_file: Path,
        source_sheet: str = "",
        source_row: int | None = None,
        fallbacks: Sequence[object] = (),
        site_position: int | None = None,
        site_residue: str = "",
        modified_sequence: str = "",
    ) -> int:
        inserted = 0
        for mapping in self.mappings(dataset_id, original_id, fallbacks):
            if not mapping.msu_locus:
                continue
            self.connection.execute(
                """INSERT INTO abundance_profiles(
                  dataset_id,original_id,msu_locus,msu_model,rap_gene,rap_model,feature_type,
                  site_position,site_residue,modified_sequence,unit,values_json,source_file,source_sheet,source_row
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    dataset_id, mapping.original_id, mapping.msu_locus, mapping.msu_model,
                    mapping.rap_gene, mapping.rap_model, feature_type, site_position, site_residue,
                    modified_sequence, unit, json_values(values), str(source_file), source_sheet, source_row,
                ),
            )
            inserted += 1
        return inserted

    def add_differential(
        self,
        dataset_id: str,
        comparison_id: str,
        original_id: object,
        *,
        feature_type: str,
        log2fc: object,
        source_file: Path,
        source_sheet: str = "",
        source_row: int | None = None,
        ratio: object = None,
        pvalue: object = None,
        padj: object = None,
        regulated: object = "",
        annotation: object = "",
        protein_accession: object = "",
        site_position: object = None,
        site_residue: object = "",
        modified_sequence: object = "",
        localization_probability: object = None,
        descriptive: bool = False,
        fallbacks: Sequence[object] = (),
    ) -> int:
        inserted = 0
        for mapping in self.mappings(dataset_id, original_id, fallbacks):
            if not mapping.msu_locus:
                continue
            self.connection.execute(
                """INSERT INTO differential_results(
                  dataset_id,comparison_id,original_id,msu_locus,msu_model,rap_gene,rap_model,feature_type,
                  log2fc,ratio,pvalue,padj,regulated,annotation,protein_accession,site_position,site_residue,
                  modified_sequence,localization_probability,descriptive,source_file,source_sheet,source_row
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    dataset_id, comparison_id, mapping.original_id, mapping.msu_locus, mapping.msu_model,
                    mapping.rap_gene, mapping.rap_model, feature_type, as_float(log2fc), as_float(ratio),
                    as_float(pvalue), as_float(padj), clean_text(regulated, 100), clean_text(annotation),
                    clean_text(protein_accession, 500), int(site_position) if as_float(site_position) is not None else None,
                    clean_text(site_residue, 10), clean_text(modified_sequence, 1_000),
                    as_float(localization_probability), int(descriptive), str(source_file), source_sheet, source_row,
                ),
            )
            inserted += 1
        return inserted


def iter_worksheet(path: Path, sheet: str, *, min_row: int = 1, max_rows: int | None = None) -> Iterator[tuple[int, tuple[object, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet]
    for offset, row in enumerate(worksheet.iter_rows(min_row=min_row, values_only=True), start=0):
        if max_rows is not None and offset >= max_rows:
            break
        yield min_row + offset, row
    workbook.close()


def register_exclusions(builder: OmicsDatabaseBuilder) -> None:
    rows = [
        dict(dataset_id="gray_planthopper_absent", display_name="灰飞虱处理", category="昆虫", assay="status", host_background="", treatment="灰飞虱", control_group="", inclusion_status="absent", inclusion_reason="未找到符合野生型同背景处理 vs 对照的已分析结果", source_root="", notes="APP显示：暂无合格数据"),
        dict(dataset_id="rbsdv_absent", display_name="RBSDV处理", category="病毒", assay="status", host_background="", treatment="RBSDV", control_group="", inclusion_status="absent", inclusion_reason="未找到符合首版边界的已分析结果", source_root="", notes="APP显示：暂无合格数据"),
        dict(dataset_id="zoujing_rsv_srna", display_name="NPB–RSV small RNA/miRNA", category="病毒", assay="small RNA", host_background="NPB", treatment="RSV", control_group="NPB", inclusion_status="candidate", inclusion_reason="small RNA不属于首版四类核心组学", source_root="/Volumes/Zoujing", notes="已登记，留待后续版本"),
        dict(dataset_id="zoujing_sp_interactions", display_name="RSV SP Y2H/IP-MS互作", category="病毒互作", assay="Y2H/IP-MS", host_background="NPB/构建体", treatment="RSV SP", control_group="", inclusion_status="excluded", inclusion_reason="互作层不属于首版，且His-SP/Myc-SP/Y2H不属于野生型处理 vs 对照", source_root="/Volumes/Zoujing", notes="Supplementary Data 2–4"),
        dict(dataset_id="zoujing_mirna_oe_mimic", display_name="miRNA OE/MIMIC", category="病毒", assay="miRNA perturbation", host_background="OE/MIMIC", treatment="RSV", control_group="", inclusion_status="excluded", inclusion_reason="OE/MIMIC不属于野生型背景", source_root="/Volumes/Zoujing", notes="首版排除"),
        dict(dataset_id="raw_only_virus_multiomics", display_name="仅有Raw data的病毒多组学", category="病毒", assay="raw", host_background="", treatment="", control_group="", inclusion_status="excluded", inclusion_reason="本版不从FASTQ或质谱Raw data重新分析", source_root="/Volumes/FAFU/原始数据", notes="只登记，不导入"),
        dict(dataset_id="mutant_oe_cross_genotype", display_name="突变体/OE/RNAi/回补/跨基因型比较", category="过滤规则", assay="mixed", host_background="mutant/OE/RNAi", treatment="", control_group="", inclusion_status="excluded", inclusion_reason="不符合野生型同背景处理 vs 对照", source_root="多个项目", notes="包括B33_BPH vs ZH11_BPH、OsPTM OE/CRISPR/trunc及相关比较"),
    ]
    for row in rows:
        builder.add_dataset(**row)


def import_leafhopper_whiteback(builder: OmicsDatabaseBuilder) -> None:
    root = Path("/Volumes/FAFU/原始数据/2025-03-11    接虫转录组    （白背  电光叶蝉）/接虫（白背，电光）测序")
    dataset_id = "npb_whiteback_electric_leafhopper_rnaseq"
    matrix = root / "Result/03.Exp/All_sample_FPKM.xls"
    builder.add_dataset(
        dataset_id=dataset_id, display_name="NPB–白背飞虱/电光叶蝉转录组", category="昆虫",
        assay="mRNA", host_background="NPB", treatment="白背飞虱；电光叶蝉",
        control_group="NPB_S；NPB_L", inclusion_status="included",
        inclusion_reason="同一NPB背景内处理 vs 匹配对照；每组4个生物学重复",
        source_root=str(root), replicate_note="每组4个生物学重复", id_namespace="MSU7 locus",
        abundance_unit="FPKM", time_order="12 → 24 → 3d",
        notes="保留原样本代码；12/24时间单位未在已核文件中明确，不作推断",
    )
    builder.add_file(dataset_id, matrix, "abundance_matrix")
    frame = pd.read_csv(matrix, sep="\t", nrows=builder.row_limit)
    sample_columns = [column for column in frame.columns if column != "Gene"]
    sample_rows = []
    for sample in sample_columns:
        group, replicate = sample.rsplit("_", 1)
        treatment = "电光叶蝉" if group.startswith("Rd") else "白背飞虱" if group.startswith("W") else "Mock/对照"
        role = "control" if group.startswith("NPB") else "treated"
        time_label = "3d" if "3d" in group else "24" if "24" in group else "12" if "12" in group else "control"
        order = {"control": 0, "12": 1, "24": 2, "3d": 3}[time_label]
        sample_rows.append(dict(sample_id=sample, original_sample_code=sample, group_id=group, condition_role=role, treatment=treatment, time_label=time_label, time_order=order, replicate=int(replicate), unit="FPKM"))
    builder.add_samples(dataset_id, sample_rows)
    for row_number, row in enumerate(frame.itertuples(index=False), start=2):
        builder.add_profile(dataset_id, row[0], row[1:], feature_type="mRNA", unit="FPKM", source_file=matrix, source_row=row_number)
    comparisons = [
        ("Rd12_L", "NPB_L", "电光叶蝉", "12", 1), ("Rd24_L", "NPB_L", "电光叶蝉", "24", 2),
        ("Rd3d_L", "NPB_L", "电光叶蝉", "3d", 3), ("W12_S", "NPB_S", "白背飞虱", "12", 1),
        ("W24_S", "NPB_S", "白背飞虱", "24", 2), ("W3d_S", "NPB_S", "白背飞虱", "3d", 3),
    ]
    for treated, control, treatment, time_label, time_order in comparisons:
        path = root / f"Result/04.DEG/{treated}_vs_{control}.DEG_all.anno.xls"
        comparison_id = f"{dataset_id}:{treated}_vs_{control}"
        builder.add_file(dataset_id, path, "differential_table")
        builder.add_comparison(
            comparison_id=comparison_id, dataset_id=dataset_id, display_name=f"{treated} vs {control}",
            treatment_group=treated, control_group=control, treatment=treatment, time_label=time_label,
            time_order=time_order, n_treatment=4, n_control=4, direction="treatment/control",
            effect_metric="source logFC", p_metric="PValue; FDR", source_file=str(path),
            notes="原项目顺序12→24→3d；不推断12/24单位",
        )
        deg = pd.read_csv(path, sep="\t", nrows=builder.row_limit)
        treatment_columns = [f"{treated}_{index}" for index in range(1, 5)]
        control_columns = [f"{control}_{index}" for index in range(1, 5)]
        derived = ((deg[treatment_columns].mean(axis=1) + 1e-9) / (deg[control_columns].mean(axis=1) + 1e-9)).map(math.log2)
        add_direction_qc(builder, dataset_id, f"{treated}_vs_{control}", deg["logFC"], derived)
        for row_number, row in enumerate(deg.itertuples(index=False), start=2):
            data = row._asdict()
            annotation = data.get("NR") or data.get("GeneName") or data.get("GeneInfo") or ""
            builder.add_differential(
                dataset_id, comparison_id, data["Gene"], feature_type="mRNA", log2fc=data.get("logFC"),
                pvalue=data.get("PValue"), padj=data.get("FDR"), annotation=annotation,
                source_file=path, source_row=row_number,
            )


def import_npb_virus_rnaseq(builder: OmicsDatabaseBuilder) -> None:
    root = Path("/Volumes/Lvshaoyuan/BCAT课题/BCAT课题   转录组-sRNA组分析/2024-02-01    BCAT新测 水稻  飞虱  转录组/lsy20231213  rice RNA-Seq")
    dataset_id = "npb_rgsv_rrsv_rnaseq"
    matrix = root / "all_fpkm.csv"
    builder.add_dataset(
        dataset_id=dataset_id, display_name="NPB–RGSV/RRSV转录组", category="病毒", assay="mRNA",
        host_background="NPB", treatment="RGSV；RRSV", control_group="NPBmock",
        inclusion_status="included", inclusion_reason="同一NPB背景内病毒处理 vs Mock；每组2个生物学重复",
        source_root=str(root), replicate_note="每组2个生物学重复", id_namespace="MSU locus/项目预测ID",
        abundance_unit="FPKM", notes="Kasa/MSG/ZCY未因品种背景未核实而纳入正式结果",
    )
    builder.add_file(dataset_id, matrix, "abundance_matrix")
    frame = pd.read_csv(matrix, nrows=builder.row_limit)
    sample_columns = ["NPBmock_1", "NPBmock_2", "NPBRG_1", "NPBRG_2", "NPBRR_1", "NPBRR_2"]
    builder.add_samples(dataset_id, [
        dict(sample_id=sample, original_sample_code=sample, group_id=sample.rsplit("_", 1)[0], condition_role="control" if "mock" in sample else "treated", treatment="Mock" if "mock" in sample else "RGSV" if "RG" in sample else "RRSV", replicate=int(sample.rsplit("_", 1)[1]), unit="FPKM")
        for sample in sample_columns
    ])
    for row_number, row in enumerate(frame[["gene_id", *sample_columns]].itertuples(index=False), start=2):
        builder.add_profile(dataset_id, row[0], row[1:], feature_type="mRNA", unit="FPKM", source_file=matrix, source_row=row_number)
    for short, treated, filename in [
        ("RGSV", "NPBRG", "NPBMockvsNPBRGSV_diff.csv"),
        ("RRSV", "NPBRR", "NPBMockvsNPBRRSV_diff.csv"),
    ]:
        path = root / "diff" / filename
        comparison_id = f"{dataset_id}:{short}_vs_mock"
        builder.add_file(dataset_id, path, "differential_table")
        builder.add_comparison(
            comparison_id=comparison_id, dataset_id=dataset_id, display_name=f"NPB {short} vs NPB Mock",
            treatment_group=treated, control_group="NPBmock", treatment=short, n_treatment=2, n_control=2,
            direction="treatment/control", effect_metric="source log2FoldChange", p_metric="pvalue; padj",
            source_file=str(path), notes="方向按文件列与处理/对照样本顺序显式固定",
        )
        deg = pd.read_csv(path, nrows=builder.row_limit)
        treated_columns = [column for column in deg.columns if column.startswith(treated + "_")]
        control_columns = [column for column in deg.columns if column.startswith("NPBmock_")]
        derived = ((deg[treated_columns].mean(axis=1) + 1e-9) / (deg[control_columns].mean(axis=1) + 1e-9)).map(math.log2)
        add_direction_qc(builder, dataset_id, f"{short}_vs_mock", deg["log2FoldChange"], derived)
        for row_number, row in enumerate(deg.itertuples(index=False), start=2):
            data = row._asdict()
            builder.add_differential(
                dataset_id, comparison_id, data["gene_id"], feature_type="mRNA",
                log2fc=data.get("log2FoldChange"), pvalue=data.get("pvalue"), padj=data.get("padj"),
                regulated=data.get("expression"), annotation=data.get("anno_info"), source_file=path, source_row=row_number,
            )

    gan_root = Path("/Volumes/Ganpeng/帅老师数据处理/20240914不同病毒基因上下调")
    for virus in ("RGSV", "RRSV", "SRBSDV"):
        path = gan_root / f"NPB_{virus}.csv"
        treatment_column = {"RGSV": "GN", "RRSV": "RN", "SRBSDV": "SN"}[virus]
        summary_id = f"npb_{virus.casefold()}_rnaseq_summary_ganpeng"
        builder.add_dataset(
            dataset_id=summary_id, display_name=f"NPB–{virus}转录组（Ganpeng现成汇总表）",
            category="病毒", assay="mRNA", host_background="NPB", treatment=virus, control_group="NPB",
            inclusion_status="included", inclusion_reason="已分析差异表；仅有组均值列，按描述性结果展示",
            source_root=str(gan_root), replicate_note="源表未提供逐重复样本", descriptive=1,
            id_namespace="MSU locus", abundance_unit="source mean normalized expression",
            notes=f"MN/{treatment_column}原列名保留；logFC与log2({treatment_column}/MN)核验方向。RGSV/RRSV与吕绍元项目相关性低，按独立实验保留",
        )
        builder.add_file(summary_id, path, "differential_summary")
        builder.add_samples(summary_id, [
            dict(sample_id="MN", original_sample_code="MN", group_id="NPB", condition_role="control", treatment="Mock", replicate=0, unit="source mean normalized expression"),
            dict(sample_id=treatment_column, original_sample_code=treatment_column, group_id=virus, condition_role="treated", treatment=virus, replicate=0, unit="source mean normalized expression"),
        ])
        comparison_id = f"{summary_id}:{virus}_vs_NPB"
        builder.add_comparison(
            comparison_id=comparison_id, dataset_id=summary_id, display_name=f"NPB {virus} vs NPB",
            treatment_group=treatment_column, control_group="MN", treatment=virus, direction="treatment/control",
            effect_metric="source logFC", p_metric="PValue; FDR", descriptive=1, source_file=str(path),
            notes="无逐重复列，标记为描述性结果",
        )
        frame = pd.read_csv(path, nrows=builder.row_limit)
        derived = ((pd.to_numeric(frame[treatment_column], errors="coerce") + 1e-9) / (pd.to_numeric(frame["MN"], errors="coerce") + 1e-9)).map(math.log2)
        add_direction_qc(builder, summary_id, f"{virus}_{treatment_column}_vs_MN", frame["logFC"], derived)
        for row_number, row in enumerate(frame.itertuples(index=False), start=2):
            data = row._asdict()
            gene = row[0]
            builder.add_profile(summary_id, gene, [data.get("MN"), data.get(treatment_column)], feature_type="mRNA", unit="source mean normalized expression", source_file=path, source_row=row_number)
            builder.add_differential(
                summary_id, comparison_id, gene, feature_type="mRNA", log2fc=data.get("logFC"),
                pvalue=data.get("PValue"), padj=data.get("FDR"), regulated=data.get("regulated"),
                source_file=path, source_row=row_number, descriptive=True,
            )
        if virus in {"RGSV", "RRSV"}:
            left_path = root / "diff" / ("NPBMockvsNPBRGSV_diff.csv" if virus == "RGSV" else "NPBMockvsNPBRRSV_diff.csv")
            left = pd.read_csv(left_path, usecols=["gene_id", "log2FoldChange"])
            right = pd.read_csv(path)
            right = right.rename(columns={right.columns[0]: "gene_id"})
            joined = left.merge(right[["gene_id", "logFC"]], on="gene_id", how="inner").dropna()
            correlation = float(joined["log2FoldChange"].corr(joined["logFC"])) if len(joined) > 2 else None
            builder.add_qc(summary_id, "cross_project_log2fc_pearson", correlation, status="info", notes=f"matched genes={len(joined)}; low concordance means not treated as a duplicate")


def import_zh11_rnaseq(builder: OmicsDatabaseBuilder) -> None:
    rgsv_root = Path("/Users/zhangshuai/NAS920同步文档/07-Wu Lab 课题/2025届----唐甜鑫-膜结合转录因子/2024-05-28    RNAseq_OsPTM")
    dataset_id = "zh11_rgsv_rnaseq"
    count_path = rgsv_root / "ZH11_RGSV_vs_ZH11_Mock.count"
    deg_path = rgsv_root / "ZH11_RGSV_vs_ZH11_Mock_all.csv"
    builder.add_dataset(
        dataset_id=dataset_id, display_name="ZH11–RGSV转录组", category="病毒", assay="mRNA",
        host_background="ZH11", treatment="RGSV", control_group="ZH11_Mock", inclusion_status="included",
        inclusion_reason="同一ZH11野生型背景内RGSV vs Mock；3个生物学重复",
        source_root=str(rgsv_root), replicate_note="每组3个生物学重复", id_namespace="MSU7 locus",
        abundance_unit="featureCounts count", notes="OE、CRISPR、截短体及跨基因型比较全部排除",
    )
    builder.add_file(dataset_id, count_path, "count_matrix")
    builder.add_file(dataset_id, deg_path, "differential_table")
    counts = pd.read_csv(count_path, nrows=builder.row_limit)
    sample_columns = list(counts.columns[1:])
    builder.add_samples(dataset_id, [
        dict(sample_id=sample, original_sample_code=sample, group_id="ZH11_Mock" if "Mock" in sample else "ZH11_RGSV", condition_role="control" if "Mock" in sample else "treated", treatment="Mock" if "Mock" in sample else "RGSV", replicate=index % 3 + 1, unit="featureCounts count")
        for index, sample in enumerate(sample_columns)
    ])
    for row_number, row in enumerate(counts.itertuples(index=False), start=2):
        builder.add_profile(dataset_id, row[0], row[1:], feature_type="mRNA", unit="featureCounts count", source_file=count_path, source_row=row_number)
    comparison_id = f"{dataset_id}:RGSV_vs_Mock"
    builder.add_comparison(
        comparison_id=comparison_id, dataset_id=dataset_id, display_name="ZH11 RGSV vs ZH11 Mock",
        treatment_group="ZH11_RGSV", control_group="ZH11_Mock", treatment="RGSV", n_treatment=3, n_control=3,
        direction="treatment/control", effect_metric="source log2FoldChange", p_metric="pvalue; padj",
        source_file=str(deg_path), notes="源表无表头；固定解析前13列，后续逗号内容合并为注释",
    )
    with deg_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        source_effects: list[float] = []
        derived_effects: list[float] = []
        for index, row in enumerate(reader, start=1):
            if builder.row_limit is not None and index > builder.row_limit:
                break
            if len(row) < 13:
                continue
            source_value = as_float(row[2])
            control_mean = pd.Series([as_float(value) for value in row[7:10]], dtype=float).mean()
            treatment_mean = pd.Series([as_float(value) for value in row[10:13]], dtype=float).mean()
            if source_value is not None and pd.notna(control_mean) and pd.notna(treatment_mean):
                source_effects.append(source_value)
                derived_effects.append(math.log2((treatment_mean + 1e-9) / (control_mean + 1e-9)))
            builder.add_differential(
                dataset_id, comparison_id, row[0], feature_type="mRNA", log2fc=row[2],
                pvalue=row[5], padj=row[6], annotation=", ".join(item.strip() for item in row[13:] if item.strip()),
                source_file=deg_path, source_row=index,
            )
        add_direction_qc(builder, dataset_id, "RGSV_vs_Mock", source_effects, derived_effects)

    rsv_root = Path("/Volumes/Zhihong Zhu/PTM课题/PTM 转录组   朱志鸿测序  AJS2260120194-G699")
    dataset_id = "zh11_rsv_rnaseq"
    matrix_path = rsv_root / "数据分析/gene_tpm_name.xlsx"
    deg_path = rsv_root / "数据分析/DE  E-VS-A    ZH11 RSV---ZH11.xlsx"
    builder.add_dataset(
        dataset_id=dataset_id, display_name="ZH11–RSV转录组", category="病毒", assay="mRNA",
        host_background="ZH11", treatment="RSV", control_group="ZH11_MocK", inclusion_status="included",
        inclusion_reason="同一ZH11野生型背景内RSV vs Mock；3个生物学重复",
        source_root=str(rsv_root), replicate_note="每组3个生物学重复", id_namespace="MSU locus",
        abundance_unit="TPM", notes="OE、KO、TM及跨基因型比较全部排除；保留源代码ZH11_MocK",
    )
    builder.add_file(dataset_id, matrix_path, "abundance_matrix")
    builder.add_file(dataset_id, deg_path, "differential_table")
    samples = ["ZH11_MocK_1", "ZH11_MocK_2", "ZH11_MocK_3", "ZH11_RSV_1", "ZH11_RSV_2", "ZH11_RSV_3"]
    builder.add_samples(dataset_id, [
        dict(sample_id=sample, original_sample_code=sample, group_id="ZH11_MocK" if "MocK" in sample else "ZH11_RSV", condition_role="control" if "MocK" in sample else "treated", treatment="Mock" if "MocK" in sample else "RSV", replicate=int(sample.rsplit("_", 1)[1]), unit="TPM")
        for sample in samples
    ])
    matrix_rows = iter_worksheet(matrix_path, "gene_tpm_name", min_row=2, max_rows=builder.row_limit)
    header = next(iter_worksheet(matrix_path, "gene_tpm_name", min_row=1, max_rows=1))[1]
    header_index = {clean_text(value): index for index, value in enumerate(header)}
    for row_number, row in matrix_rows:
        builder.add_profile(dataset_id, row[0], [row[header_index[sample]] for sample in samples], feature_type="mRNA", unit="TPM", source_file=matrix_path, source_sheet="gene_tpm_name", source_row=row_number, fallbacks=[row[1]])
    sheet = "DE  E-VS-A    ZH11 RSV---ZH11"
    rows = iter_worksheet(deg_path, sheet, min_row=1, max_rows=(builder.row_limit + 1) if builder.row_limit else None)
    _, header = next(rows)
    index = {clean_text(value): position for position, value in enumerate(header)}
    comparison_id = f"{dataset_id}:RSV_vs_Mock"
    builder.add_comparison(
        comparison_id=comparison_id, dataset_id=dataset_id, display_name="ZH11 RSV vs ZH11 Mock",
        treatment_group="ZH11_RSV", control_group="ZH11_MocK", treatment="RSV", n_treatment=3, n_control=3,
        direction="treatment/control", effect_metric="source log2FoldChange", p_metric="pvalue; padj",
        source_file=str(deg_path), source_sheet=sheet,
    )
    source_effects: list[float] = []
    derived_effects: list[float] = []
    for row_number, row in rows:
        source_value = as_float(row[index["log2FoldChange"]])
        treatment_mean = pd.Series([as_float(row[index[name]]) for name in ("ZH11_RSV_1", "ZH11_RSV_2", "ZH11_RSV_3")], dtype=float).mean()
        control_mean = pd.Series([as_float(row[index[name]]) for name in ("ZH11_MocK_1", "ZH11_MocK_2", "ZH11_MocK_3")], dtype=float).mean()
        if source_value is not None and pd.notna(treatment_mean) and pd.notna(control_mean):
            source_effects.append(source_value)
            derived_effects.append(math.log2((treatment_mean + 1e-9) / (control_mean + 1e-9)))
        builder.add_differential(
            dataset_id, comparison_id, row[index["Gene"]], feature_type="mRNA",
            log2fc=row[index["log2FoldChange"]], pvalue=row[index["pvalue"]], padj=row[index["padj"]],
            source_file=deg_path, source_sheet=sheet, source_row=row_number,
        )
    add_direction_qc(builder, dataset_id, "RSV_vs_Mock", source_effects, derived_effects)


def import_bph_rnaseq(builder: OmicsDatabaseBuilder) -> None:
    root = Path("/Users/zhangshuai/NAS920同步文档/16-投稿中论文/2026-03-25    BPH33  论文/组学/转录组分析")
    specs = [
        ("6h", root / "感 6h vs 0h.xls", ["J01", "J02", "J03"], ["J04", "J05", "J06"]),
        ("24h", root / "感 24h vs 0h-版本1.xls", ["J01", "J02", "J03"], ["J07", "J08", "J09"]),
    ]
    for order, (time_label, path, controls, treated) in enumerate(specs, start=1):
        dataset_id = f"bph_susceptible_rnaseq_{time_label}"
        builder.add_dataset(
            dataset_id=dataset_id, display_name=f"感性背景–褐飞虱 {time_label} 转录组", category="昆虫",
            assay="mRNA", host_background="感性背景（源文件标签：感）", treatment=f"褐飞虱 {time_label}",
            control_group="0h", inclusion_status="included",
            inclusion_reason="明确感性背景内褐飞虱处理 vs 0h；3个生物学重复",
            source_root=str(root), replicate_note="每组3个生物学重复", id_namespace="RAP gene + MSU locus",
            abundance_unit="FPKM", time_order=f"0h → {time_label}", notes="不纳入抗性背景及跨基因型比较",
        )
        builder.add_file(dataset_id, path, "differential_table_with_FPKM")
        frame = pd.read_csv(path, sep="\t", nrows=builder.row_limit)
        samples = [*controls, *treated]
        builder.add_samples(dataset_id, [
            dict(sample_id=sample, original_sample_code=sample, group_id="0h" if sample in controls else time_label, condition_role="control" if sample in controls else "treated", treatment="Mock/0h" if sample in controls else "褐飞虱", time_label="0h" if sample in controls else time_label, time_order=0 if sample in controls else order, replicate=(controls.index(sample) + 1) if sample in controls else (treated.index(sample) + 1), unit="FPKM")
            for sample in samples
        ])
        comparison_id = f"{dataset_id}:{time_label}_vs_0h"
        builder.add_comparison(
            comparison_id=comparison_id, dataset_id=dataset_id, display_name=f"感性背景 褐飞虱 {time_label} vs 0h",
            treatment_group=time_label, control_group="0h", treatment="褐飞虱", time_label=time_label,
            time_order=order, n_treatment=3, n_control=3, direction="treatment/control",
            effect_metric="source log2FC", p_metric="FDR", source_file=str(path),
        )
        treatment_columns = [f"{sample}_FPKM" for sample in treated]
        control_columns = [f"{sample}_FPKM" for sample in controls]
        derived = ((frame[treatment_columns].mean(axis=1) + 1e-9) / (frame[control_columns].mean(axis=1) + 1e-9)).map(math.log2)
        add_direction_qc(builder, dataset_id, f"{time_label}_vs_0h", frame["log2FC"], derived)
        for row_number, row in enumerate(frame.itertuples(index=False), start=2):
            data = row._asdict()
            gene = data.get("_0") or row[0]
            fallback = data.get("gene_name", "")
            values = [data.get(f"{sample}_FPKM") for sample in samples]
            builder.add_profile(dataset_id, gene, values, feature_type="mRNA", unit="FPKM", source_file=path, source_row=row_number, fallbacks=[fallback])
            builder.add_differential(
                dataset_id, comparison_id, gene, feature_type="mRNA", log2fc=data.get("log2FC"),
                padj=data.get("FDR"), regulated=data.get("regulated"), annotation=data.get("NR_annotation"),
                source_file=path, source_row=row_number, fallbacks=[fallback],
            )


def _header_positions(header: Sequence[object]) -> dict[str, list[int]]:
    positions: dict[str, list[int]] = {}
    for index, value in enumerate(header):
        key = clean_text(value)
        if key:
            positions.setdefault(key, []).append(index)
    return positions


def import_ms_project(
    builder: OmicsDatabaseBuilder,
    *,
    dataset_id: str,
    display_name: str,
    root: Path,
    report_folder: str,
    assay: str,
    groups: Sequence[str],
    comparisons: Sequence[tuple[str, str, str]],
    background: str,
    category: str,
    excluded_groups: str,
) -> None:
    report_root = root / report_folder
    identified = report_root / "2-Basic_analysis/MS_identified_information.xlsx"
    diff_path = report_root / "4-Differentially_expressed_protein/T-test_analysis/Differentially_expressed_statistics.xlsx"
    feature_type = {"total proteome": "protein", "phosphoproteome": "phosphosite", "ubiquitome": "ubiquitination_site"}[assay]
    builder.add_dataset(
        dataset_id=dataset_id, display_name=display_name, category=category, assay=assay,
        host_background=background, treatment="；".join(treatment for treatment, _, _ in comparisons),
        control_group=comparisons[0][1], inclusion_status="included",
        inclusion_reason="同一野生型背景处理 vs 匹配对照；使用公司已分析定量与差异表",
        source_root=str(report_root), replicate_note="每组单个汇总定量，无可核实生物学重复",
        descriptive=1, id_namespace="MSU protein model/项目预测蛋白ID",
        abundance_unit="normalized protein/site quantitation", notes=f"描述性结果；排除：{excluded_groups}",
    )
    builder.add_file(dataset_id, identified, "identified_quantification")
    builder.add_file(dataset_id, diff_path, "differential_statistics")
    builder.add_samples(dataset_id, [
        dict(sample_id=group, original_sample_code=group, group_id=group, condition_role="control" if group == comparisons[0][1] else "treated", treatment="Mock/对照" if group == comparisons[0][1] else group, replicate=1, unit="normalized protein/site quantitation")
        for group in groups
    ])
    sheet = "Protein_quant" if assay == "total proteome" else "Site_quant Normalized"
    header_row = 1 if assay == "total proteome" else 2
    rows = iter_worksheet(identified, sheet, min_row=header_row, max_rows=(builder.row_limit + 1) if builder.row_limit else None)
    _, header = next(rows)
    positions = _header_positions(header)
    group_indexes = {group: positions[group][-1] for group in groups if group in positions}
    accession_index = positions["Protein accession"][0]
    description_index = positions.get("Protein description", [None])[0]
    site_index = positions.get("Position", [None])[0]
    residue_index = positions.get("Amino acid", [None])[0]
    modified_index = positions.get("Modified sequence", [None])[0]
    for row_number, row in rows:
        accession = row[accession_index]
        if not clean_text(accession):
            continue
        builder.add_profile(
            dataset_id, accession, [row[group_indexes[group]] if group in group_indexes else None for group in groups],
            feature_type=feature_type, unit="normalized protein/site quantitation", source_file=identified,
            source_sheet=sheet, source_row=row_number,
            site_position=int(row[site_index]) if site_index is not None and as_float(row[site_index]) is not None else None,
            site_residue=clean_text(row[residue_index], 10) if residue_index is not None else "",
            modified_sequence=clean_text(row[modified_index], 1_000) if modified_index is not None else "",
            fallbacks=[row[description_index]] if description_index is not None else (),
        )
    workbook = load_workbook(diff_path, read_only=True, data_only=True)
    for treatment, control, sheet_name in comparisons:
        if sheet_name not in workbook.sheetnames:
            builder.add_qc(dataset_id, "missing_comparison_sheet", text=sheet_name, status="fail")
            continue
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        positions = _header_positions(header)
        ratio_label = next((key for key in positions if key.endswith(" Ratio") and treatment in key and control in key), "")
        if not ratio_label:
            ratio_label = next((key for key in positions if key.endswith(" Ratio")), "")
        comparison_id = f"{dataset_id}:{treatment}_vs_{control}"
        builder.add_comparison(
            comparison_id=comparison_id, dataset_id=dataset_id, display_name=f"{treatment} vs {control}",
            treatment_group=treatment, control_group=control, treatment=treatment,
            n_treatment=1, n_control=1, direction="treatment/control", effect_metric="log2(source ratio)",
            p_metric="not provided", descriptive=1, source_file=str(diff_path), source_sheet=sheet_name,
            notes="公司阈值筛选的差异条目；无生物学重复，按描述性结果展示",
        )
        for offset, row in enumerate(rows, start=2):
            if builder.row_limit is not None and offset > builder.row_limit + 1:
                break
            accession = row[positions["Protein accession"][0]]
            ratio = as_float(row[positions[ratio_label][0]]) if ratio_label else None
            if not clean_text(accession) or ratio is None or ratio <= 0:
                continue
            builder.add_differential(
                dataset_id, comparison_id, accession, feature_type=feature_type, log2fc=math.log2(ratio), ratio=ratio,
                regulated=row[positions.get("Regulated Type", [0])[0]] if "Regulated Type" in positions else "",
                annotation=row[positions.get("Protein description", [0])[0]] if "Protein description" in positions else "",
                protein_accession=accession,
                site_position=row[positions["Position"][0]] if "Position" in positions else None,
                site_residue=row[positions["Amino acid"][0]] if "Amino acid" in positions else "",
                modified_sequence=row[positions["Modified sequence"][0]] if "Modified sequence" in positions else "",
                localization_probability=row[positions["Localization probability"][0]] if "Localization probability" in positions else None,
                descriptive=True, source_file=diff_path, source_sheet=sheet_name, source_row=offset,
            )
    workbook.close()


def import_mass_spectrometry(builder: OmicsDatabaseBuilder) -> None:
    virus_root = Path("/Users/zhangshuai/NAS920同步文档/07-Wu Lab 课题/2024届----刘鸿飞    病毒侵染水稻表型/2023-09  病毒蛋白组---修饰组数据")
    virus_groups = ["NPB", "RRSV", "RGSV", "SRBSDV", "RGDV"]
    virus_comparisons = [(virus, "NPB", f"{virus}vsNPB") for virus in virus_groups[1:]]
    for suffix, folder, assay in [
        ("proteome", "cloud_report_48114   蛋白质组学", "total proteome"),
        ("phosphoproteome", "cloud_report_48132    磷酸化修饰组学", "phosphoproteome"),
        ("ubiquitome", "cloud_report_48123  泛素化修饰组学", "ubiquitome"),
    ]:
        import_ms_project(
            builder, dataset_id=f"npb_virus_{suffix}", display_name=f"NPB病毒处理{assay}",
            root=virus_root, report_folder=folder, assay=assay, groups=virus_groups,
            comparisons=virus_comparisons, background="NPB", category="病毒",
            excluded_groups="无；仅纳入NPB、RRSV、RGSV、SRBSDV、RGDV",
        )

    bph_root = Path("/Users/zhangshuai/NAS920同步文档/16-投稿中论文/2026-03-25    BPH33  论文/组学/景杰生物  蛋白组-磷酸化-泛素化/公司结题报告")
    bph_groups = ["ZH11_M", "ZH11_BPH"]
    bph_comparisons = [("ZH11_BPH", "ZH11_M", "ZH11_BPHvsZH11_M")]
    for suffix, folder, assay in [
        ("proteome", "cloud_report_24809", "total proteome"),
        ("phosphoproteome", "cloud_report_24827", "phosphoproteome"),
        ("ubiquitome", "cloud_report_24818", "ubiquitome"),
    ]:
        import_ms_project(
            builder, dataset_id=f"zh11_bph_{suffix}", display_name=f"ZH11褐飞虱处理{assay}",
            root=bph_root, report_folder=folder, assay=assay, groups=bph_groups,
            comparisons=bph_comparisons, background="ZH11", category="昆虫",
            excluded_groups="B33_M、B33_BPH以及B33_BPH vs ZH11_BPH跨基因型比较",
        )


def convert_rsv_microarray(source: Path, output_dir: Path) -> Path:
    converted = output_dir / "RSV-microarray.xlsx"
    if converted.exists():
        return converted
    output_dir.mkdir(parents=True, exist_ok=True)
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if not executable:
        raise RuntimeError("LibreOffice/soffice unavailable for legacy .xls conversion")
    completed = subprocess.run(
        [executable, "--headless", "--convert-to", "xlsx", "--outdir", str(output_dir), str(source)],
        check=False, capture_output=True, text=True, timeout=180,
    )
    if completed.returncode != 0 or not converted.is_file():
        raise RuntimeError(f"RSV microarray conversion failed: {completed.stderr.strip() or completed.stdout.strip()}")
    return converted


def import_historical_microarray(builder: OmicsDatabaseBuilder, processed_dir: Path) -> None:
    source = Path("/Volumes/FAFU/原始数据/2022-03-24    水稻病毒芯片分析结果-马伟分析/水稻病毒芯片分析/RSV-microarray.xls")
    dataset_id = "historical_rsv_microarray"
    builder.add_dataset(
        dataset_id=dataset_id, display_name="RSV历史芯片", category="历史芯片", assay="microarray",
        host_background="源表未明确标注", treatment="RSV 3D/6D/9D/12D", control_group="matched mock",
        inclusion_status="included", inclusion_reason="目标病毒RSV的已分析历史芯片；作为独立历史层展示",
        source_root=str(source.parent), replicate_note="源表仅给Median Log2 signal intensity，重复数未恢复",
        descriptive=1, historical=1, id_namespace="MSU locus + probe", abundance_unit="Median Log2 signal intensity",
        time_order="3D → 6D → 9D → 12D", notes="不与RNA-seq/蛋白组原始数值直接比较",
    )
    builder.add_file(dataset_id, source, "legacy_microarray_workbook")
    converted = convert_rsv_microarray(source, processed_dir)
    builder.add_file(dataset_id, converted, "derived_readable_copy", notes="由LibreOffice转换；原.xls未修改")
    samples = []
    for order, time_label in enumerate(("3D", "6D", "9D", "12D"), start=1):
        samples.extend([
            dict(sample_id=f"RSV{time_label}", original_sample_code=f"RSV{time_label}", group_id=f"RSV_{time_label}", condition_role="treated", treatment="RSV", time_label=time_label, time_order=order, replicate=0, unit="Median Log2 signal intensity"),
            dict(sample_id=f"mock{time_label}", original_sample_code=f"mock{time_label}", group_id=f"mock_{time_label}", condition_role="control", treatment="Mock", time_label=time_label, time_order=order, replicate=0, unit="Median Log2 signal intensity"),
        ])
    builder.add_samples(dataset_id, samples)
    for order, time_label in enumerate(("3D", "6D", "9D", "12D"), start=1):
        builder.add_comparison(
            comparison_id=f"{dataset_id}:RSV_{time_label}_vs_mock", dataset_id=dataset_id,
            display_name=f"RSV {time_label} vs matched mock", treatment_group=f"RSV_{time_label}",
            control_group=f"mock_{time_label}", treatment="RSV", time_label=time_label, time_order=order,
            direction="treatment/control", effect_metric="source Log2 Ratio", p_metric="source significance encoded as numeric vs ns",
            descriptive=1, source_file=str(source), source_sheet="SupTableS1",
        )
    rows = iter_worksheet(converted, "SupTableS1", min_row=5, max_rows=builder.row_limit)
    for row_number, row in rows:
        probe, locus, annotation = row[1], row[2], row[3]
        signals = row[4:12]
        builder.add_profile(dataset_id, probe, signals, feature_type="microarray_probe", unit="Median Log2 signal intensity", source_file=source, source_sheet="SupTableS1", source_row=row_number, fallbacks=[locus])
        for index, time_label in enumerate(("3D", "6D", "9D", "12D")):
            log2fc = as_float(row[12 + index])
            if log2fc is None:
                continue
            builder.add_differential(
                dataset_id, f"{dataset_id}:RSV_{time_label}_vs_mock", probe,
                feature_type="microarray_probe", log2fc=log2fc, annotation=annotation,
                descriptive=True, source_file=source, source_sheet="SupTableS1", source_row=row_number,
                fallbacks=[locus],
            )


def register_zoujing_files(builder: OmicsDatabaseBuilder) -> None:
    root = Path("/Volumes/Zoujing/相分离SP-SE 文章投稿/Nature Commucinations/2025-06-14   原则性接收   返稿文件")
    for number, dataset_id, role in [
        (2, "zoujing_sp_interactions", "Y2H_interactors"),
        (3, "zoujing_sp_interactions", "His-SP_IP-MS"),
        (4, "zoujing_sp_interactions", "Myc-SP_IP-MS"),
        (5, "zoujing_rsv_srna", "small_RNA_miRNA"),
    ]:
        builder.add_file(dataset_id, root / f"Supplementary Data {number}.xlsx", role)


DATA_DICTIONARY = [
    ("datasets", "dataset_id", "稳定数据集ID；APP与导出使用"),
    ("datasets", "inclusion_status", "included/excluded/candidate/absent"),
    ("datasets", "descriptive", "1表示无可核实生物学重复，仅作描述性展示"),
    ("samples", "original_sample_code", "原项目样本代码，原样保留"),
    ("comparisons", "direction", "统一为treatment/control；正log2FC表示处理上调"),
    ("gene_mappings", "msu_locus", "APP主键；去除MSU model后缀"),
    ("gene_mappings", "msu_model", "MSU蛋白/转录本model；一对多逐条保留"),
    ("gene_mappings", "rap_gene", "RAP gene ID"),
    ("gene_mappings", "rap_model", "RAP transcript/model；源数据存在时保留"),
    ("gene_mappings", "mapping_status", "mapped/multi_model/multi_locus/unmapped"),
    ("differential_results", "log2fc", "源分析表log2FC；蛋白/PTM仅对源ratio取log2"),
    ("differential_results", "descriptive", "无重复或重复数不明时为1"),
    ("abundance_profiles", "values_json", "按samples.sample_order排列的已有FPKM/TPM/count/归一化定量"),
    ("abundance_profiles", "feature_type", "mRNA/protein/phosphosite/ubiquitination_site/microarray_probe"),
    ("files", "sha256", "源文件指纹，用于重复检查和可追溯核验"),
]


def export_registry(db_path: Path, output_path: Path) -> None:
    connection = sqlite3.connect(db_path)
    queries = {
        "Datasets": "SELECT * FROM datasets ORDER BY inclusion_status, category, assay, dataset_id",
        "Samples": "SELECT * FROM samples ORDER BY dataset_id, sample_order",
        "Comparisons": "SELECT * FROM comparisons ORDER BY dataset_id, time_order, comparison_id",
        "Files": "SELECT * FROM files ORDER BY dataset_id, role, path",
        "Excluded_Candidates": "SELECT * FROM datasets WHERE inclusion_status <> 'included' ORDER BY inclusion_status,dataset_id",
        "Duplicate_Files": "SELECT * FROM files WHERE duplicate_of <> '' ORDER BY dataset_id,path",
        "ID_Mapping_QC": """
            SELECT dataset_id, mapping_status, COUNT(*) AS mapping_rows,
                   COUNT(DISTINCT original_id) AS original_ids,
                   COUNT(DISTINCT CASE WHEN msu_locus <> '' THEN msu_locus END) AS msu_loci
            FROM gene_mappings GROUP BY dataset_id,mapping_status ORDER BY dataset_id,mapping_status
        """,
        "ID_Mapping_Exceptions": """
            SELECT dataset_id,original_id,msu_locus,msu_model,rap_gene,rap_model,mapping_status,mapping_note
            FROM gene_mappings WHERE mapping_status <> 'mapped' ORDER BY dataset_id,mapping_status,original_id
        """,
        "QC_Summary": "SELECT * FROM qc_metrics ORDER BY status,dataset_id,metric",
    }
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True, name="Arial")
    for sheet_name, query in queries.items():
        frame = pd.read_sql_query(query, connection)
        worksheet = workbook.create_sheet(sheet_name[:31])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max(len(frame.columns), 1))}{max(len(frame) + 1, 1)}"
        for column_index, column in enumerate(frame.columns, start=1):
            cell = worksheet.cell(1, column_index, column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(row_index, column_index, "" if value is None else value)
        for column_index, column in enumerate(frame.columns, start=1):
            sample = [len(str(column))] + [len(str(value)) for value in frame[column].head(300) if value is not None]
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(sample, default=10) + 2, 48)
    worksheet = workbook.create_sheet("Data_Dictionary")
    worksheet.append(["table", "field", "description"])
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in DATA_DICTIONARY:
        worksheet.append(row)
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 28
    worksheet.column_dimensions["C"].width = 80
    workbook.save(output_path)
    connection.close()


def validate_database(db_path: Path) -> list[tuple[str, str, float | None]]:
    connection = sqlite3.connect(db_path)
    checks: list[tuple[str, str, float | None]] = []
    integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
    checks.append(("sqlite_integrity", "pass" if integrity == "ok" else "fail", None))
    cross_genotype = connection.execute(
        "SELECT COUNT(*) FROM comparisons WHERE treatment_group LIKE 'B33%' OR control_group LIKE 'B33%'"
    ).fetchone()[0]
    checks.append(("no_B33_comparisons", "pass" if cross_genotype == 0 else "fail", float(cross_genotype)))
    forbidden = connection.execute(
        """SELECT COUNT(*) FROM comparisons
           WHERE lower(display_name) LIKE '%oe%' OR lower(display_name) LIKE '%crispr%'
              OR lower(display_name) LIKE '%rnai%' OR lower(display_name) LIKE '%trunc%'"""
    ).fetchone()[0]
    checks.append(("no_mutant_OE_RNAi_comparisons", "pass" if forbidden == 0 else "fail", float(forbidden)))
    for treatment, expected in (("白背飞虱", 3), ("电光叶蝉", 3)):
        count = connection.execute("SELECT COUNT(*) FROM comparisons WHERE treatment=?", (treatment,)).fetchone()[0]
        checks.append((f"{treatment}_comparisons", "pass" if count == expected else "fail", float(count)))
    unmapped = connection.execute("SELECT COUNT(DISTINCT original_id) FROM gene_mappings WHERE mapping_status='unmapped'").fetchone()[0]
    total = connection.execute("SELECT COUNT(DISTINCT dataset_id || '|' || original_id) FROM gene_mappings").fetchone()[0]
    rate = 1.0 - unmapped / max(total, 1)
    checks.append(("overall_mapping_rate", "pass" if rate >= 0.75 else "review", rate))
    connection.close()
    return checks


def write_validation_report(path: Path, checks: Sequence[tuple[str, str, float | None]], db_path: Path) -> None:
    lines = [
        "# Wu Lab analysed-omics database validation",
        "",
        f"- Database: `{db_path}`",
        "- Raw FASTQ / MS Raw reanalysis: **No**",
        "- Primary key: MSU locus without model suffix",
        "",
        "| Check | Status | Value |",
        "|---|---:|---:|",
    ]
    for name, status, value in checks:
        formatted = "" if value is None else f"{value:.6g}"
        lines.append(f"| {name} | {status} | {formatted} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--mapping", type=Path, default=DEFAULT_MAPPING)
    parser.add_argument("--dry-run", action="store_true", help="read at most 100 source rows into an isolated preview")
    parser.add_argument("--force", action="store_true", help="replace an existing generated database/registry")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    output_dir = args.output_dir / "dry_run" if args.dry_run else args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    db_path = output_dir / "wulab_omics_v1.sqlite"
    registry_path = output_dir / "dataset_registry.xlsx"
    validation_path = output_dir / "validation_report.md"
    generated = [db_path, registry_path, validation_path]
    existing = [path for path in generated if path.exists()]
    if existing and not args.force:
        raise FileExistsError("Refusing to overwrite generated files without --force: " + ", ".join(map(str, existing)))
    for path in existing:
        path.unlink()

    resolver = IdentifierResolver(args.mapping)
    builder = OmicsDatabaseBuilder(db_path, resolver, row_limit=100 if args.dry_run else None)
    try:
        register_exclusions(builder)
        register_zoujing_files(builder)
        import_leafhopper_whiteback(builder)
        import_npb_virus_rnaseq(builder)
        import_zh11_rnaseq(builder)
        import_bph_rnaseq(builder)
        import_mass_spectrometry(builder)
        import_historical_microarray(builder, output_dir / "processed_sources")
        builder.connection.commit()
    finally:
        builder.close()
    export_registry(db_path, registry_path)
    checks = validate_database(db_path)
    write_validation_report(validation_path, checks, db_path)
    failed = [name for name, status, _ in checks if status == "fail"]
    print(f"Database: {db_path}")
    print(f"Registry: {registry_path}")
    print(f"Validation: {validation_path}")
    print(f"Checks: {len(checks) - len(failed)}/{len(checks)} non-failing")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
