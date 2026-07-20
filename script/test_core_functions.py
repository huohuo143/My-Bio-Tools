#!/usr/bin/env python3
"""Unit tests for the data-processing logic behind the Streamlit pages."""

from __future__ import annotations

import gzip
import io
import inspect
import os
from pathlib import Path
import random
import sys
import tempfile
import threading
import time
import unittest
import zipfile
import xml.etree.ElementTree as ET

os.environ.setdefault("ARROW_DEFAULT_MEMORY_POOL", "system")

ROOT = Path(__file__).resolve().parent.parent
APP_SOURCE = ROOT / "app_source"
sys.path.insert(0, str(APP_SOURCE))

from RAP_MSU_convert import convert_gene_ids, detect_id_type  # noqa: E402
from analysis_jobs import AnalysisJobManager, RiceGeneAnalysisRequest  # noqa: E402
import RiceData_crawler as ricedata  # noqa: E402
from RiceData_crawler import parse_detail_html, parse_reference_html  # noqa: E402
from RGAP_sequence_downloader import (  # noqa: E402
    build_download_zip,
    canonicalize_msu_id,
    fetch_rgap_sequence,
    parse_rgap_ids,
    parse_rgap_sequence_html,
)
from extract_fasta import extract_fasta_records  # noqa: E402
from fasta_rename import decode_upload, parse_mapping_details, rename_fasta_with_stats  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from PIL import Image  # noqa: E402
from prediction_services import run_nlstradamus  # noqa: E402
from prediction_visualization import build_prediction_chart_artifacts  # noqa: E402
from primer_design import design_primer_pairs, normalize_dna_sequence, validate_dna_sequence  # noqa: E402
from report_builder import CHINESE_FONT, build_report_artifacts  # noqa: E402
from protein_domain_analysis import parse_matches_payload, build_domain_artifacts  # noqa: E402
from gene_structure_analysis import parse_gene_model, build_gene_structure_artifacts  # noqa: E402
from promoter_regulation_analysis import build_tfbs_artifacts, parse_plantregmap_html  # noqa: E402
from variation_analysis import build_variation_artifacts, parse_vcf  # noqa: E402
from mirna_rnai_analysis import parse_psrnatarget_html  # noqa: E402
from methods_guide import method_entries  # noqa: E402
from literature_evidence_analysis import build_query, import_manual_evidence, fetch_rapdb_genetic_evidence, genetic_evidence_from_ricedata  # noqa: E402
import rice_efp as efp_module  # noqa: E402
import rice_gene_analysis as rice_gene_analysis_module  # noqa: E402
from sequence_visualization import build_sequence_relationship_artifacts  # noqa: E402
from rice_gene_core import (  # noqa: E402
    AnalysisBundle,
    CDS,
    PredictionProviderAttempt,
    PredictionRegion,
    PredictionResult,
    SequenceRecord,
    build_reference_sequence_indexes,
    normalize_cds,
    normalize_protein,
    sequence_digest,
    translate_cds,
)
from rice_efp import (  # noqa: E402
    EfpExpressionRecord,
    build_efp_chart_artifacts,
    duplicate_expression_count,
    efp_source_display_label,
    expression_top_rows,
    fetch_efp_records,
    parse_efp_result_html,
    parse_expression_table_html,
    unique_expression_records,
)
from rice_seq_extractor import extract_bundled_sequences, query_prefix, record_matches  # noqa: E402
from rice_utr_promoter_downloader import (  # noqa: E402
    FIVE_UTR,
    PROMOTER,
    THREE_UTR,
    TRANSCRIPT_SCOPE_ALL,
    RiceSequenceResult,
    build_download_zip as build_utr_download_zip,
    extract_utr_sequences,
    fetch_gene_payload as fetch_utr_gene_payload,
    format_fasta as format_utr_fasta,
    parse_input_ids as parse_utr_input_ids,
    promoter_region,
    rap_gene_from_transcript,
    resolve_input_ids as resolve_utr_input_ids,
)
from tool_a import analyze_sequences, parse_sequences  # noqa: E402


class CoreFunctionTests(unittest.TestCase):
    class Upload:
        def __init__(self, data: bytes, name: str):
            self._data = data
            self.name = name

        def getvalue(self) -> bytes:
            return self._data

    class FakeResponse:
        def __init__(self, text: str):
            self.text = text
            self.apparent_encoding = "utf-8"
            self.encoding = "utf-8"

        def raise_for_status(self) -> None:
            return None

    class FakeRiceDataSession:
        def __init__(self):
            self.calls: list[str] = []

        def get(self, url: str, **kwargs):
            self.calls.append(url)
            if "accessions_switch" in url:
                cells = [
                    '<td style="border-bottom:1px solid silver">label</td>',
                    '<td style="border-bottom:1px solid silver">Gene Name</td>',
                    '<td style="border-bottom:1px solid silver"><em>GS</em></td>',
                    '<td style="border-bottom:1px solid silver"><a name=Os01g0100100">RAP</a></td>',
                    '<td style="border-bottom:1px solid silver"><a orf=LOC_Os01g01010.1">MSU</a></td>',
                    '<td style="border-bottom:1px solid silver"><a term=123">NCBI</a></td>',
                    '<td style="border-bottom:1px solid silver"><a href="nuccore/AK000001">cDNA</a></td>',
                    '<td style="border-bottom:1px solid silver"><a href="nuccore/NM_001"><a href="protein/NP_001"></td>',
                    '<td style="border-bottom:1px solid silver"><a href="uniprot/Q00001">UniProt</a></td>',
                ]
                return CoreFunctionTests.FakeResponse(
                    '<td height="22"><a>RID0001</a></td>' + "".join(cells)
                )
            return CoreFunctionTests.FakeResponse(
                '<td style="padding: 5px; font-size: 14px">【生物学功能】抗性相关【亚细胞定位】细胞核</td>'
            )

    def test_sequence_composition_supports_fasta_and_ambiguity(self) -> None:
        records = parse_sequences(">seq1\nACGTNN\n>seq2\nGGCC\n")
        frame = analyze_sequences(records)
        self.assertEqual(frame["Sequence ID"].tolist(), ["seq1", "seq2"])
        self.assertEqual(frame.loc[0, "Length (nt)"], 6)
        self.assertEqual(frame.loc[0, "GC (%)"], 50.0)
        self.assertEqual(frame.loc[0, "N (%)"], 33.33)
        self.assertEqual(frame.loc[1, "Reverse complement"], "GGCC")

    def test_fasta_extraction_and_missing_report(self) -> None:
        handle = io.StringIO(">seq1 desc\nACGT\n>seq2.1\nGGCC\n")
        records, missing, scanned = extract_fasta_records(
            handle, ["seq2", "absent"], ignore_version=True
        )
        self.assertEqual([record.id for record in records], ["seq2.1"])
        self.assertEqual(missing, ["absent"])
        self.assertEqual(scanned, 2)

    def test_fasta_rename_reports_unmatched_and_duplicates(self) -> None:
        parsed = parse_mapping_details("old1\tnew1\nold1\tnewest\nbadline\n")
        result = rename_fasta_with_stats(">old1 desc\nACGT\n>old2\nTTAA\n", parsed.mapping)
        self.assertIn(">newest desc", result.text)
        self.assertEqual(result.renamed_count, 1)
        self.assertEqual(result.unchanged_ids, ["old2"])
        self.assertEqual(parsed.duplicate_ids, ["old1"])
        self.assertEqual(parsed.invalid_lines, [3])

    def test_fasta_rename_reads_gzip_upload(self) -> None:
        upload = self.Upload(gzip.compress(b">seq1\nACGT\n"), "input.fa.gz")
        self.assertEqual(decode_upload(upload), ">seq1\nACGT\n")

    def test_rap_msu_mixed_conversion(self) -> None:
        rap_to_msu = {"Os01g0100100": ("LOC_Os01g01010.1", "LOC_Os01g01010.2")}
        msu_to_rap = {
            "LOC_Os01g01010": ("Os01g0100100",),
            "LOC_Os01g01010.1": ("Os01g0100100",),
        }
        result = convert_gene_ids(
            ["Os01g0100100", "LOC_Os01g01010", "bad"], rap_to_msu, msu_to_rap
        )
        self.assertEqual(result.loc[0, "mapping_count"], 2)
        self.assertEqual(result.loc[1, "converted"], "Os01g0100100")
        self.assertEqual(result.loc[2, "status"], "invalid_id")
        self.assertEqual(detect_id_type("LOC_Os01g01010.2"), "MSU")
        self.assertEqual(rap_gene_from_transcript("Os03t0106400-01"), "Os03g0106400")

    def test_rice_query_normalization_and_real_record(self) -> None:
        self.assertEqual(query_prefix("Os01g0100100"), ("Os01t0100100", False))
        self.assertTrue(record_matches("Os01t0100100-01", "Os01g0100100"))
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "tiny.fa.gz"
            with gzip.open(path, "wt") as handle:
                handle.write(">Os01t0100100-01 desc\nACGT\n>Os02t0200200-01\nTTAA\n")
            records, missing, scanned = extract_bundled_sequences(
                str(path), ("Os01g0100100", "Os03g0300300")
            )
        self.assertEqual(records[0][0], "Os01t0100100-01")
        self.assertEqual(missing, ["Os03g0300300"])
        self.assertEqual(scanned, 2)

    def test_rgap_id_parsing_and_canonicalization(self) -> None:
        self.assertEqual(canonicalize_msu_id("loc_os10G33000.1"), "LOC_Os10g33000.1")
        self.assertEqual(
            parse_rgap_ids("LOC_Os10g33000.1, loc_os10g33000.1\nLOC_Os01g01010"),
            ["LOC_Os10g33000.1", "LOC_Os01g01010"],
        )
        self.assertIsNone(canonicalize_msu_id("Os10g3300000"))

    def test_rgap_sequence_parser_and_zip_bundle(self) -> None:
        html = """
        <html><head><title>LOC_Os10g33000 Sequence Information</title></head><body>
        <p><b>LOC_Os10g33000 sequence information</b></p>
        <p><b>Genomic sequence length: </b>4 nucleotides<br>
        <b>CDS length: </b>3 nucleotides<br>
        <b>Protein length: </b>1 amino acids<br>
        <b>Putative Function: </b>test protein</p>
        <p>Genomic Sequence</p><pre>&gt;LOC_Os10g33000\nACGT</pre>
        <p>CDS</p><pre>&gt;LOC_Os10g33000.1\nATG</pre>
        <p>Protein</p><pre>&gt;LOC_Os10g33000.1\nM*</pre>
        </body></html>
        """
        record = parse_rgap_sequence_html(html, "LOC_Os10g33000.1")
        self.assertEqual(record.status, "matched")
        self.assertEqual(record.locus_id, "LOC_Os10g33000")
        self.assertEqual(record.genomic_sequence, "ACGT")
        self.assertEqual(record.cds_sequence, "ATG")
        self.assertEqual(record.protein_length, 1)
        self.assertEqual(record.putative_function, "test protein")
        self.assertEqual(record.validation_note, "")

        archive_bytes = build_download_zip(
            [record], ["Genomic Sequence", "CDS", "Protein"]
        )
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
            names = set(archive.namelist())
            self.assertIn("RGAP_genomic_sequences.fasta", names)
            self.assertIn("RGAP_CDS_sequences.fasta", names)
            self.assertIn("RGAP_protein_sequences.fasta", names)
            self.assertIn("RGAP_download_summary.csv", names)
            self.assertIn(">LOC_Os10g33000.1\nM*", archive.read("RGAP_protein_sequences.fasta").decode())

    def test_rgap_fetch_pipeline_and_missing_record(self) -> None:
        html = """
        <p><b>Genomic sequence length: </b>4 nucleotides<br>
        <b>CDS length: </b>3 nucleotides<br>
        <b>Protein length: </b>1 amino acids<br></p>
        <p>Genomic Sequence</p><pre>&gt;LOC_Os10g33000\nACGT</pre>
        <p>CDS</p><pre>&gt;LOC_Os10g33000.1\nATG</pre>
        <p>Protein</p><pre>&gt;LOC_Os10g33000.1\nM*</pre>
        """

        class FakeSession:
            def __init__(self, response_text: str):
                self.response_text = response_text
                self.calls: list[str] = []

            def get(self, url: str, **kwargs):
                self.calls.append(url)
                return CoreFunctionTests.FakeResponse(self.response_text)

        session = FakeSession(html)
        record = fetch_rgap_sequence("LOC_Os10g33000.1", session=session)
        self.assertEqual(record.status, "matched")
        self.assertEqual(len(session.calls), 1)
        self.assertIn("orf=LOC_Os10g33000.1", session.calls[0])

        missing = parse_rgap_sequence_html(
            "<title>Error: Sequence Information Not Found</title>"
            "<p>The locus or model name was not found.</p>",
            "LOC_Os10g33000.9",
        )
        self.assertEqual(missing.status, "not_found")
        invalid = fetch_rgap_sequence("Os10g3300000", session=session)
        self.assertEqual(invalid.status, "invalid_id")
        self.assertEqual(len(session.calls), 1)

    def test_utr_id_resolution_preserves_msu_one_to_many_mapping(self) -> None:
        identifiers = parse_utr_input_ids(
            "Os01g0100100, os01G0100100\nOs01t0100100-01\nLOC_Os01g01010.1\nbad"
        )
        self.assertEqual(len(identifiers), 4)
        targets = resolve_utr_input_ids(
            identifiers,
            {"LOC_Os01g01010.1": ("Os01g0100100", "Os01g0100200")},
        )
        self.assertEqual(len(targets), 5)
        self.assertEqual(targets[0].rap_gene_id, "Os01g0100100")
        self.assertEqual(targets[1].requested_transcript_id, "Os01t0100100-01")
        mapped = [item for item in targets if item.input_type == "MSU"]
        self.assertEqual([item.rap_gene_id for item in mapped], ["Os01g0100100", "Os01g0100200"])
        self.assertTrue(all(item.status == "mapped_one_to_many" for item in mapped))
        self.assertEqual(targets[-1].status, "invalid_id")

    def test_promoter_coordinates_clip_boundaries_and_keep_gene_orientation(self) -> None:
        positive = promoter_region("1", 1000, 1500, 1, 2000, {"1": 3000})
        self.assertEqual(positive, (1, 999, 1, "1:1..999:1"))
        negative = promoter_region("1", 100, 500, -1, 2000, {"1": 1200})
        self.assertEqual(negative, (501, 1200, -1, "1:501..1200:-1"))
        self.assertEqual(promoter_region("1", 5001, 6000, 1, 500)[0:2], (4501, 5000))
        self.assertEqual(promoter_region("1", 5001, 6000, 1, 4000)[0:2], (1001, 5000))
        with self.assertRaises(ValueError):
            promoter_region("1", 5001, 6000, 1, 499)
        with self.assertRaises(ValueError):
            promoter_region("1", 5001, 6000, 1, 4001)

    def test_utr_slicing_uses_transcript_utr_features(self) -> None:
        transcript = {
            "UTR": [
                {"type": "five_prime_utr", "start": 1, "end": 2},
                {"object_type": "five_prime_UTR", "start": 10, "end": 11},
                {"type": "three_prime_utr", "start": 20, "end": 22},
            ],
            "Exon": [{"start": 1, "end": 12}],
        }
        five, three, note = extract_utr_sequences(transcript, "AAAACCCCCGGG")
        self.assertEqual(five, "AAAA")
        self.assertEqual(three, "GGG")
        self.assertEqual(note, "")

    def test_utr_promoter_pipeline_and_zip_with_mocked_ensembl(self) -> None:
        class FakeJsonResponse:
            def __init__(self, payload: dict[str, object]):
                self.payload = payload

            def raise_for_status(self) -> None:
                return None

            def json(self) -> dict[str, object]:
                return self.payload

        class FakeEnsemblSession:
            def __init__(self):
                self.calls: list[tuple[str, dict[str, object]]] = []

            def get(self, url: str, **kwargs):
                params = dict(kwargs.get("params") or {})
                self.calls.append((url, params))
                if "/lookup/id/" in url:
                    return FakeJsonResponse(
                        {
                            "id": "Os01g0100100",
                            "object_type": "Gene",
                            "assembly_name": "IRGSP-1.0",
                            "source": "RAP2022-09-01",
                            "seq_region_name": "1",
                            "start": 1001,
                            "end": 2000,
                            "strand": 1,
                            "canonical_transcript": "Os01t0100100-01",
                            "Transcript": [
                                {
                                    "id": "Os01t0100100-01",
                                    "UTR": [
                                        {"type": "five_prime_utr", "start": 1001, "end": 1002},
                                        {"type": "three_prime_utr", "start": 1998, "end": 2000},
                                    ],
                                    "Exon": [{"start": 1001, "end": 1010}],
                                }
                            ],
                        }
                    )
                if "/sequence/region/" in url:
                    return FakeJsonResponse({"seq": "ACGT" * 125})
                if "/sequence/id/" in url:
                    return FakeJsonResponse({"seq": "AACCCCCGGG"})
                raise AssertionError(f"unexpected URL: {url}")

        session = FakeEnsemblSession()
        payload = fetch_utr_gene_payload(
            "Os01g0100100",
            "",
            TRANSCRIPT_SCOPE_ALL,
            (FIVE_UTR, THREE_UTR, PROMOTER),
            500,
            chromosome_lengths={"1": 5000},
            session=session,
        )
        self.assertEqual(payload.status, "matched")
        self.assertEqual(payload.annotation_source, "RAP2022-09-01")
        self.assertEqual(len(payload.promoter_sequence), 500)
        self.assertEqual(payload.transcripts[0].five_utr_sequence, "AA")
        self.assertEqual(payload.transcripts[0].three_utr_sequence, "GGG")
        target = resolve_utr_input_ids(["LOC_Os01g01010.1"], {"LOC_Os01g01010.1": ("Os01g0100100",)})[0]
        results = [RiceSequenceResult(target=target, payload=payload)]
        self.assertIn("type=5UTR", format_utr_fasta(results, FIVE_UTR))
        self.assertIn("orientation=gene_5to3", format_utr_fasta(results, PROMOTER))
        archive_bytes = build_utr_download_zip(
            results,
            (FIVE_UTR, THREE_UTR, PROMOTER),
            500,
        )
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
            names = set(archive.namelist())
            self.assertIn("rice_5UTR_sequences.fasta", names)
            self.assertIn("rice_3UTR_sequences.fasta", names)
            self.assertIn("rice_promoter_500bp_sequences.fasta", names)
            self.assertIn("rice_utr_promoter_summary.csv", names)

    def test_cds_and_protein_validation_never_guesses_reading_frame(self) -> None:
        normalized, errors = normalize_cds("ATGAA")
        self.assertEqual(normalized, "ATGAA")
        self.assertTrue(any("不是 3 的倍数" in error for error in errors))
        protein, errors = translate_cds("ATGTAGGCT")
        self.assertEqual(protein, "M*A")
        self.assertTrue(any("内部终止密码子" in error for error in errors))
        _, errors = normalize_cds("ATG!")
        self.assertTrue(any("非法字符" in error for error in errors))
        normalized_protein, errors = normalize_protein("MKRKRTKQKRRK*")
        self.assertEqual(normalized_protein, "MKRKRTKQKRRK")
        self.assertFalse(errors)

    def test_exact_reverse_indexes_preserve_multiple_matches(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "tiny.fa.gz"
            with gzip.open(path, "wt") as handle:
                handle.write(
                    ">Os01t0100100-01\nATGAAATAA\n"
                    ">Os01t0100200-01\nATGAAATAA\n"
                    ">Os01t0100300-01\nATGCCCTAA\n"
                )
            build_reference_sequence_indexes.cache_clear()
            cds_index, protein_index = build_reference_sequence_indexes(str(path))
        self.assertEqual(
            cds_index[sequence_digest("ATGAAATAA")],
            ("Os01t0100100-01", "Os01t0100200-01"),
        )
        self.assertEqual(len(protein_index[sequence_digest("MK")]), 2)
        self.assertEqual(cds_index.get(sequence_digest("ATGTTTTAA"), ()), ())
        build_reference_sequence_indexes.cache_clear()

    def test_nlstradamus_original_binary_and_parser(self) -> None:
        result = run_nlstradamus(
            "known_nls",
            "MKRKRTKQKRRKAAAAAAAAAAAAAAAAAAAAAAAAAAAA",
            model=1,
            cutoff=0.6,
        )
        self.assertEqual(
            result.status,
            "matched",
            f"NLStradamus error={result.error!r}; raw={result.raw_text!r}",
        )
        self.assertEqual(result.classification, "NLS detected")
        self.assertEqual((result.regions[0].start, result.regions[0].end), (2, 15))
        self.assertAlmostEqual(result.regions[0].score or 0, 0.954, places=3)

    def test_prediction_visualization_svg_and_600_dpi_png(self) -> None:
        predictions = [
            PredictionResult(
                protein_id="CLV3_ARATH",
                tool="SignalP 6.0",
                version="6.0",
                status="matched",
                classification="Sec/SPI",
                provider="biolib",
                probabilities={"OTHER": 0.01, "Sec/SPI": 0.99},
                regions=[PredictionRegion("signal peptide", 1, 25)],
            ),
            PredictionResult(
                protein_id="CLV3_ARATH",
                tool="DeepTMHMM 1.0",
                version="1.0",
                status="matched",
                classification="signal peptide",
                provider="biolib",
                regions=[PredictionRegion("signal", 1, 25), PredictionRegion("outside", 26, 96)],
            ),
            PredictionResult(
                protein_id="CLV3_ARATH", tool="TMHMM 2.0", version="2.0", status="matched",
                classification="no TM helix", provider="dtu_web",
            ),
            PredictionResult(
                protein_id="CLV3_ARATH", tool="TargetP 2.0", version="2.0", status="matched",
                classification="SP", provider="dtu_web", probabilities={"SP": 0.93, "OTHER": 0.07},
                regions=[PredictionRegion("targeting peptide", 1, 21)],
            ),
            PredictionResult(
                protein_id="CLV3_ARATH", tool="cNLS Mapper", version="web", status="matched",
                classification="cNLS detected", provider="nls_mapper_web",
                regions=[PredictionRegion("NLS", 62, 75, score=7.5)],
            ),
            PredictionResult(
                protein_id="CLV3_ARATH", tool="NLStradamus 1.8", version="1.8", status="matched",
                classification="NLS detected", provider="local",
                regions=[PredictionRegion("NLS", 64, 77, score=0.91)],
            ),
            PredictionResult(
                protein_id="negative_fixture", tool="SignalP 6.0", version="6.0", status="failed",
                provider="biolib", error="mock provider failure",
            ),
        ]
        charts = build_prediction_chart_artifacts(
            predictions,
            {"CLV3_ARATH": "M" * 96, "negative_fixture": "M" * 60},
        )
        self.assertIn("combined_CLV3_ARATH.svg", charts)
        self.assertTrue(charts["combined_CLV3_ARATH.svg"].lstrip().startswith(b"<?xml"))
        with Image.open(io.BytesIO(charts["combined_CLV3_ARATH.png"])) as image:
            self.assertGreaterEqual(float(image.info.get("dpi", (0, 0))[0]), 590)
        self.assertIn(b"Service unavailable", charts["combined_negative_fixture.svg"])
        self.assertNotIn(b"mock provider failure", charts["combined_negative_fixture.svg"])

    def test_report_bundle_has_fixed_sheets_fastas_and_explicit_word_fonts(self) -> None:
        bundle = AnalysisBundle(
            mode="单基因深度分析",
            input_type="RAP/MSU ID",
            inputs=["Os01g0100100"],
            mapping_rows=[
                {
                    "input_id": "Os01g0100100",
                    "input_type": "RAP gene",
                    "resolved_rap_gene": "Os01g0100100",
                    "resolved_msu_id": "LOC_Os01g01010",
                    "status": "matched",
                }
            ],
            sequences=[
                SequenceRecord(
                    input_id="Os01g0100100",
                    resolved_rap_gene="Os01g0100100",
                    transcript_id="Os01t0100100-01",
                    sequence_type=CDS,
                    sequence="ATGAAATAA",
                    source="test",
                )
            ],
            predictions=[
                PredictionResult(
                    protein_id="Os01t0100100-01",
                    tool="SignalP 6.0",
                    version="6.0",
                    status="failed",
                    error="mock timeout",
                    raw_text=">Os01t0100100-01\nMK\n",
                    provider="dtu_web",
                    provider_job_id="mock-job",
                    attempts=[PredictionProviderAttempt("dtu_web", "failed", "mock-job", "https://example/job", "mock timeout")],
                )
            ],
            ricedata_rows=[
                {
                    "check": "Os01g0100100",
                    "GeneName": "mock gene",
                    "GeneSymbol": "MG",
                    "RAP_Locus": "Os01g0100100",
                    "MSU_Locus": "LOC_Os01g01010.1",
                    "status": "matched",
                    "error": "",
                }
            ],
            efp_rows=[
                EfpExpressionRecord(
                    input_id="Os01g0100100",
                    msu_locus="LOC_Os01g01010",
                    data_source="rice_rma",
                    data_source_label="Developmental atlas (RMA)",
                    group="1",
                    tissue="Seedling Root",
                    expression_level=9.48,
                    standard_deviation=0.07,
                    samples="Root_1, Root_2, Root_3",
                    probe_id="Os.1.1.S1_at",
                )
            ],
            warnings=["部分预测失败，但报告仍生成。"],
            sources=["IRGSP-1.0"],
            generated_at="2026-07-17T12:00:00+08:00",
        )
        charts = build_efp_chart_artifacts(bundle.efp_rows)
        prediction_charts = build_prediction_chart_artifacts(bundle.predictions, {"Os01t0100100-01": "MK"})
        artifacts = build_report_artifacts(
            bundle,
            "Os01g0100100",
            efp_charts=charts,
            prediction_charts=prediction_charts,
            prediction_raw_artifacts={"dtu/mock/result.gff3": b"##gff-version 3\n"},
        )
        workbook = load_workbook(io.BytesIO(artifacts["xlsx"]), read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Overview",
                "Interpretation",
                "Mechanism Evidence",
                "AI Synthesis",
                "ID_Mapping",
                "RiceData",
                "RiceData_References",
                "eFP_Expression",
                "eFP_Top_Tissues",
                "eFP_Source_Glossary",
                "Sequence_Summary",
                "Sequence_Plot_Data",
                "Prediction_Summary",
                "Prediction_Scores",
                "Prediction_Regions",
                "Protein_Domains",
                "Functional_Sites",
                "Transcript_Models",
                "Gene_Features",
                "Promoter_TFBS",
                "Upstream_TF",
                "Variants",
                "Haplotype_Summary",
                "miRNA_Targets",
                "RNAi_Offtargets",
                "Literature",
                "Genetic_Evidence",
                "Lab_Omics_Datasets",
                "Lab_Omics_Comparisons",
                "Lab_Omics_Samples",
                "Lab_Omics_Differential",
                "Lab_Omics_Profiles",
                "Lab_Omics_Status",
                "Omics_Published_Evidence",
                "Omics_Consensus_Scores",
                "Omics_QC",
                "Omics_Dataset_Context",
                "Omics_Dataset_Registry",
                "Warnings_Sources",
            ],
        )
        glossary_headers = [cell.value for cell in next(workbook["eFP_Source_Glossary"].iter_rows())]
        self.assertTrue({"name_zh", "reference", "replicate_note"}.issubset(glossary_headers))
        with zipfile.ZipFile(io.BytesIO(artifacts["zip"])) as archive:
            names = set(archive.namelist())
            fasta_names = {name for name in names if name.startswith("sequences/") and name.endswith(".fasta")}
            self.assertEqual(len(fasta_names), 6)
            self.assertTrue(any(name.endswith("_error.txt") for name in names))
            self.assertIn("manifest.json", names)
            self.assertIn("interpretation/interpretation.csv", names)
            self.assertIn("interpretation/interpretation.json", names)
            self.assertIn("interpretation/status.json", names)
            self.assertIn("annotations/ricedata_gene_annotations.csv", names)
            self.assertIn("annotations/ricedata_references.csv", names)
            self.assertIn("expression/efp_expression_values.csv", names)
            self.assertTrue(any(name.endswith(".svg") for name in names if name.startswith("expression/figures/")))
            self.assertIn("predictions/provider_trace.json", names)
            self.assertIn("predictions/raw/dtu/mock/result.gff3", names)
            self.assertIn("predictions/figures/combined_Os01t0100100-01.svg", names)
            self.assertIn("protein_domains/protein_domains.csv", names)
            self.assertIn("gene_structure/transcript_models.csv", names)
            self.assertIn("promoter_regulation/promoter_tfbs.csv", names)
            self.assertIn("variation/variants.csv", names)
            self.assertIn("mirna_rnai/mirna_targets.csv", names)
            self.assertIn("literature_evidence/literature.csv", names)
            self.assertIn("literature_evidence/ricedata_references.csv", names)
            self.assertIn("sequences/sequence_relationship_plot_data.csv", names)
            png_name = next(name for name in names if name.startswith("expression/figures/") and name.endswith(".png"))
            with Image.open(io.BytesIO(archive.read(png_name))) as image:
                self.assertGreaterEqual(float(image.info.get("dpi", (0, 0))[0]), 590)
        w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        required_parts = ["word/document.xml", "word/header1.xml", "word/footer1.xml"]
        with zipfile.ZipFile(io.BytesIO(artifacts["docx"])) as archive:
            for name in required_parts:
                root = ET.fromstring(archive.read(name))
                fonts = list(root.iter(f"{{{w}}}rFonts"))
                self.assertTrue(fonts, name)
                for font in fonts:
                    self.assertEqual(font.get(f"{{{w}}}eastAsia"), CHINESE_FONT)
                    self.assertEqual(font.get(f"{{{w}}}ascii"), "Times New Roman")
                    self.assertEqual(font.get(f"{{{w}}}hAnsi"), "Times New Roman")
                    self.assertEqual(font.get(f"{{{w}}}cs"), "Times New Roman")
            styles_root = ET.fromstring(archive.read("word/styles.xml"))
            required_styles = {"Normal", "Title", "Heading1", "Heading2", "Heading3", "ListBullet", "ListNumber"}
            checked_styles = set()
            for style in styles_root.iter(f"{{{w}}}style"):
                style_id = style.get(f"{{{w}}}styleId")
                if style_id not in required_styles:
                    continue
                fonts = list(style.iter(f"{{{w}}}rFonts"))
                self.assertTrue(fonts, style_id)
                for font in fonts:
                    self.assertEqual(font.get(f"{{{w}}}eastAsia"), CHINESE_FONT)
                    self.assertEqual(font.get(f"{{{w}}}ascii"), "Times New Roman")
                    self.assertEqual(font.get(f"{{{w}}}hAnsi"), "Times New Roman")
                checked_styles.add(style_id)
            self.assertEqual(checked_styles, required_styles)
            document_text = "".join(ET.fromstring(archive.read("word/document.xml")).itertext())
            self.assertIn("2. 已知功能、突变体与关联文献", document_text)
            self.assertIn("4. 序列、转录本与蛋白结构", document_text)
            self.assertIn("A1.2 eFP 官方来源与重复/汇总结构", document_text)
            self.assertNotIn("12. 文献与已知遗传证据", document_text)

    def test_v160_domain_gene_structure_and_negative_strand_fixtures(self) -> None:
        domains, sites = parse_matches_payload({"matches": [{"signature": {"accession": "PF0001", "name": "Kinase", "signatureLibraryRelease": {"library": "Pfam", "version": "36"}}, "locations": [{"start": 10, "end": 80}, {"start": 70, "end": 120, "sites": [{"description": "active site", "siteLocations": [{"start": 75, "end": 75, "residue": "D"}]}]}]}]}, "P1", 150, "fixture://interpro")
        self.assertEqual(len(domains), 2); self.assertEqual(sites[0]["start"], 75)
        self.assertTrue({"protein_domains_P1.svg", "protein_domains_P1.pdf", "protein_domains_P1.png"}.issubset(build_domain_artifacts(domains, sites)))
        gene = {"id": "Os01g0100100", "assembly_name": "IRGSP-1.0", "seq_region_name": "1", "start": 100, "end": 500, "strand": -1, "canonical_transcript": "Os01t0100100-01", "Transcript": [{"id": "Os01t0100100-01", "start": 100, "end": 500, "strand": -1, "is_canonical": 1, "biotype": "protein_coding", "Exon": [{"id": "E1", "start": 400, "end": 500}, {"id": "E2", "start": 100, "end": 200}]}]}
        transcripts, features = parse_gene_model(gene, "input", "全部 transcript", "fixture://ensembl")
        self.assertEqual(transcripts[0]["strand"], -1); self.assertEqual(features[0]["feature_id"], "E1")
        self.assertTrue(any(name.endswith(".png") for name in build_gene_structure_artifacts(transcripts, features)))

    def test_v171_ricedata_reference_and_evidence_mapping(self) -> None:
        detail = parse_detail_html(
            '<td style="padding: 5px; font-size: 14px">'
            '【突变体表型】osbcat1 异亮氨酸下降（Sun et al. 2020）。'
            '<a href="../../reference/papers.aspx?id=67133">paper</a>'
            '<a href="../../reference/papers.aspx?id=72287">paper</a></td>'
        )
        self.assertEqual([item["reference_id"] for item in detail["reference_links"]], ["67133", "72287"])
        parsed = parse_reference_html(
            '<h2>Natural variation in the OsbZIP18 promoter contributes to branched-chain amino acid levels in rice</h2>'
            '<div>DOI: 10.1111/nph.16800 PMID: 32654152 Year: 2020</div>',
            "67133",
            "https://www.ricedata.cn/reference/papers.aspx?id=67133",
        )
        self.assertEqual(parsed["doi"], "10.1111/nph.16800")
        self.assertEqual(parsed["pmid"], "32654152")
        rows = [{
            "check": "Os03g0106400", "GeneID": "52496", "GeneSymbol": "OsBCAT1",
            "RAP_Locus": "Os03g0106400", "MSU_Locus": "LOC_Os03g01600",
            "突变体表型": "突变体 osbcat1 中异亮氨酸显著下降（Sun et al. 2020）。",
        }]
        references = [
            {**parsed, "gene_id": "52496", "authors": "Sun et al."},
            {"reference_id": "72287", "gene_id": "52496", "title": "OsBCAT2 salt tolerance", "doi": "10.1111/nph.19551", "year": "2024", "authors": "Wang et al."},
        ]
        evidence = genetic_evidence_from_ricedata(rows, references)
        self.assertEqual(evidence[0]["linked_dois"], "10.1111/nph.16800")
        self.assertIn("citation_year", evidence[0]["matched_by"])
        self.assertTrue(evidence[0]["verification_status"].startswith("直接支持"))

    def test_v171_single_cell_efp_submits_rap_gene(self) -> None:
        class Session:
            def __init__(self):
                self.posts = []

            def post(self, url, data, timeout):
                self.posts.append(data)
                return CoreFunctionTests.FakeResponse('<a href="../output/test.html">table</a>')

            def get(self, url, timeout):
                return CoreFunctionTests.FakeResponse('<table><tr><td>Group</td><td>Cell type</td><td>12.5</td><td>0.4</td></tr></table>')

        session = Session()
        original = efp_module.get_session
        efp_module.get_session = lambda: session
        try:
            records = fetch_efp_records("Os03t0106400-01", "LOC_Os03g01600", "rice_single_cell", rap_locus="Os03g0106400")
            self.assertEqual(session.posts[0]["primaryGene"], "Os03g0106400")
            self.assertEqual(records[0].submitted_id, "Os03g0106400")
            self.assertEqual(records[0].id_namespace, "RAP")
        finally:
            efp_module.get_session = original

    def test_v197_sequence_relationship_formats_and_result_tabs(self) -> None:
        bundle = AnalysisBundle(
            mode="单基因深度分析",
            input_type="RAP/MSU ID",
            inputs=["Os03t0106400-01"],
            mapping_rows=[{"input_id": "Os03t0106400-01", "resolved_rap_gene": "Os03g0106400", "resolved_msu_id": "LOC_Os03g01600", "status": "matched"}],
            sequences=[
                SequenceRecord("Os03t0106400-01", "Os03g0106400", "LOC_Os03g01600", "Os03t0106400-01", CDS, "ATG" * 10, "fixture CDS", "IRGSP-1.0"),
                SequenceRecord("Os03t0106400-01", "Os03g0106400", "LOC_Os03g01600", "Os03t0106400-01", "Protein", "M" * 10, "fixture protein", "IRGSP-1.0"),
            ],
        )
        rows, artifacts, csv_bytes = build_sequence_relationship_artifacts(bundle)
        self.assertEqual({Path(name).suffix for name in artifacts}, {".svg", ".pdf", ".png"})
        self.assertIn("translation_consistency", rows[0])
        self.assertTrue(csv_bytes.startswith(b"\xef\xbb\xbf"))
        source = inspect.getsource(rice_gene_analysis_module._show_results)
        for label in ("总览", "功能与证据", "表达", "序列与结构", "调控与变异", "AI 深度解读", "结论与来源"):
            self.assertIn(f'"{label}"', source)

    def test_v160_plantregmap_vcf_and_ref_conflict_fixtures(self) -> None:
        html = '<script>const x={"data":[["OsTF##M001","bZIP","tx1","91-99","+","1e-6","2e-5","ACGT"]]};</script>'
        hits, candidates = parse_plantregmap_html(html, input_id="g", rap_gene="Os01g0100100", transcript_id="tx1", promoter_length=100, pvalue=1e-4)
        self.assertEqual(hits[0]["relative_start"], -10); self.assertEqual(candidates[0]["rank"], 1)
        self.assertIn("tfbs_tx1.png", build_tfbs_artifacts(hits))
        vcf = b"##fileformat=VCFv4.2\n##reference=IRGSP-1.0\n#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tS1\tS2\tS3\n1\t102\tv1\tG\tA\t.\tPASS\tANN=A|missense_variant|MODERATE|G|g|transcript|tx|protein_coding|1/1|c.2G>A|p.Gly1Asp\tGT\t0/0\t0/1\t1/1\n1\t103\tv2\tT\tC\t.\tPASS\t.\tGT\t0/0\t0/1\t0/1\n"
        variants, haplotypes, warnings = parse_vcf(vcf, "x.vcf", input_id="g", rap_gene="Os01g0100100", transcript_id="tx", gene_start=100, gene_end=110, strand=1, features=[{"feature_type": "CDS", "start": 100, "end": 110}], reference_sequence="AAGAAAAAAA")
        self.assertEqual(variants[0]["ref_validation"], "matched"); self.assertEqual(variants[1]["status"], "ref_mismatch")
        self.assertTrue(haplotypes); self.assertFalse(any("VCF 无样本" in warning for warning in warnings))
        variation_charts = build_variation_artifacts(variants, haplotypes)
        self.assertIn("variant_overview.png", variation_charts)
        self.assertIn("haplotype_frequency.png", variation_charts)

    def test_v160_mirna_and_manual_evidence_fixtures(self) -> None:
        html = "<table><tr><th>miRNA</th><th>Target</th><th>Expectation</th><th>UPE</th></tr><tr><td>osa-miR1</td><td>Os01t1</td><td>2.0</td><td>18.2</td></tr></table>"
        rows = parse_psrnatarget_html(html, "fixture://psrna", "g")
        self.assertEqual(rows[0]["small_rna"], "osa-miR1"); self.assertEqual(rows[0]["evidence_status"], "计算预测")
        query = build_query(["Os01g0100100", "LOC_Os01g01010"])
        self.assertIn("Oryza sativa", query)
        imported = import_manual_evidence(b"rap_gene,evidence_type\nOs01g0100100,knockout\n", "evidence.csv")
        self.assertEqual(imported[0]["source_type"], "manual_import")
        class RapResponse:
            text = '<table><tr><th>Allelic variation</th><td>mutant phenotype</td></tr></table>'
            content = text.encode()
            url = "fixture://rapdb"
            def raise_for_status(self): return None
        class RapSession:
            def get(self, *args, **kwargs):
                return RapResponse()
        evidence, raw, warnings = fetch_rapdb_genetic_evidence(["Os01g0100100"], session=RapSession())
        self.assertEqual(evidence[0]["source_type"], "RAP-DB"); self.assertTrue(raw); self.assertFalse(warnings)

    def test_primer_sequence_normalization(self) -> None:
        sequence = normalize_dna_sequence(">gene\nacgt acgt\nNN\n")
        self.assertEqual(sequence, "ACGTACGTNN")
        self.assertIn("序列短于 60 nt，难以设计常规 PCR 引物", validate_dna_sequence(sequence))

    def test_primer3_returns_paired_results(self) -> None:
        random.seed(42)
        sequence = "".join(random.choice("ACGT") for _ in range(700))
        frame, result = design_primer_pairs(
            "test_gene",
            sequence,
            primer_num=3,
            min_size=18,
            opt_size=20,
            max_size=25,
            min_tm=57,
            opt_tm=60,
            max_tm=63,
            min_gc=40,
            max_gc=60,
            product_min=250,
            product_max=650,
            forward_start=0,
            forward_len=180,
            reverse_start=500,
            reverse_len=180,
        )
        self.assertEqual(len(frame), 3)
        self.assertEqual(result["PRIMER_PAIR_NUM_RETURNED"], 3)
        self.assertTrue(frame["Forward sequence (5'-3')"].str.len().ge(18).all())

    def test_ricedata_detail_parser(self) -> None:
        html = '<td style="padding: 5px; font-size: 14px">【生物学功能】抗性相关【亚细胞定位】细胞核</td>'
        result = parse_detail_html(html)
        self.assertEqual(result["生物学功能"], "抗性相关")
        self.assertEqual(result["亚细胞定位"], "细胞核")

    def test_ricedata_fetch_pipeline_with_mocked_site(self) -> None:
        previous_session = getattr(ricedata._thread_local, "session", None)
        fake_session = self.FakeRiceDataSession()
        ricedata._thread_local.session = fake_session
        try:
            rows = ricedata.fetch_gene_records("Os01g0100100")
        finally:
            if previous_session is None:
                delattr(ricedata._thread_local, "session")
            else:
                ricedata._thread_local.session = previous_session
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["status"], "matched")
        self.assertEqual(rows[0]["GeneID"], "RID0001")
        self.assertEqual(rows[0]["RAP_Locus"], "Os01g0100100")
        self.assertEqual(rows[0]["MSU_Locus"], "LOC_Os01g01010.1")
        self.assertEqual(rows[0]["生物学功能"], "抗性相关")
        self.assertEqual(len(fake_session.calls), 2)

    def test_ricedata_fast_mode_skips_detail_request(self) -> None:
        previous_session = getattr(ricedata._thread_local, "session", None)
        fake_session = self.FakeRiceDataSession()
        ricedata._thread_local.session = fake_session
        try:
            rows = ricedata.fetch_gene_records("Os01g0100100", include_details=False)
        finally:
            if previous_session is None:
                delattr(ricedata._thread_local, "session")
            else:
                ricedata._thread_local.session = previous_session
        self.assertEqual(rows[0]["status"], "matched")
        self.assertEqual(rows[0]["RAP_Locus"], "Os01g0100100")
        self.assertEqual(rows[0]["生物学功能"], "")
        self.assertEqual(len(fake_session.calls), 1)

    def test_ricedata_cache_reuses_only_successful_results(self) -> None:
        previous_session = getattr(ricedata._thread_local, "session", None)
        fake_session = self.FakeRiceDataSession()
        ricedata._thread_local.session = fake_session
        ricedata._result_cache.clear()
        try:
            first = ricedata.cached_fetch_gene_records("Os01g0100100", include_details=False)
            second = ricedata.cached_fetch_gene_records("Os01g0100100", include_details=False)
        finally:
            ricedata._result_cache.clear()
            if previous_session is None:
                delattr(ricedata._thread_local, "session")
            else:
                ricedata._thread_local.session = previous_session
        self.assertEqual(first, second)
        self.assertIsNot(first, second)
        self.assertEqual(len(fake_session.calls), 1)

    def test_efp_result_and_expression_table_parsers(self) -> None:
        result_html = """
        <a href="../output/efp-test123.html">table</a>
        <li>Os.21356.1.S1_at was used as the probe set identifier for your primary gene</li>
        """
        table_url, probe_id = parse_efp_result_html(result_html)
        self.assertEqual(
            table_url,
            "https://bar.utoronto.ca/transcriptomics/efp_rice/output/efp-test123.html",
        )
        self.assertEqual(probe_id, "Os.21356.1.S1_at")
        table_html = """
        <table>
          <tr><th>Group</th><th>Tissue</th><th>Expression</th><th>SD</th><th>Samples</th><th>Links</th></tr>
          <tr><td>1</td><td>Seedling Root</td><td>9.48</td><td>0.07</td><td>R1, R2, R3</td>
              <td><a href="https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE6893">GEO</a></td></tr>
          <tr><td>1</td><td>Seedling Root</td><td>9.48</td><td>0.07</td><td>R1, R2, R3</td>
              <td><a href="https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE6893">GEO</a></td></tr>
        </table>
        """
        records = parse_expression_table_html(
            table_html,
            input_id="Os01g0100100",
            msu_locus="LOC_Os01g01010",
            data_source="rice_rma",
            probe_id=probe_id,
        )
        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].tissue, "Seedling Root")
        self.assertEqual(records[0].expression_level, 9.48)
        self.assertEqual(records[0].standard_deviation, 0.07)
        self.assertEqual(records[0].probe_id, probe_id)
        self.assertEqual(duplicate_expression_count(records), 1)
        self.assertEqual(len(unique_expression_records(records)), 1)
        self.assertEqual(len(expression_top_rows(records)), 1)
        self.assertIn("发育组织图谱", efp_source_display_label("rice_rma"))
        self.assertTrue(
            all(
                all(info.get(field) for field in ("name_zh", "reference", "replicate_note"))
                for info in efp_module.EFP_SOURCE_GLOSSARY.values()
            )
        )

    def test_methods_guide_covers_all_modules_and_efp_sources(self) -> None:
        entries = method_entries()
        self.assertEqual(len(entries), 39)
        self.assertEqual(len({entry["entry_id"] for entry in entries}), 39)
        self.assertEqual(sum(entry["section"] == "eFP 数据源" for entry in entries), 12)
        required = ("section", "module", "data_nature", "inputs", "source", "method", "outputs", "boundary")
        self.assertTrue(all(all(str(entry.get(field, "")).strip() for field in required) for entry in entries))

    @staticmethod
    def _job_request(name: str) -> RiceGeneAnalysisRequest:
        return RiceGeneAnalysisRequest(
            project_name=name,
            mode="单基因深度分析",
            input_type="RAP/MSU ID",
            text="Os01g0100100",
            selected_types=("CDS",),
            promoter_length=2000,
            transcript_scope=TRANSCRIPT_SCOPE_ALL,
            selected_predictors=(),
            signalp_mode="fast",
            cnls_cutoff=5.0,
            nlstradamus_model=1,
            nlstradamus_cutoff=0.6,
            max_workers=1,
            include_ricedata=False,
            include_efp=False,
        )

    def test_background_job_queue_cancel_and_isolation(self) -> None:
        manager = AnalysisJobManager(max_history=5)
        gate = threading.Event()

        def runner(request, reporter):
            reporter.complete("mapping", "mapped")
            while not gate.wait(0.01):
                reporter.check_cancel()
            reporter.complete("sequences", "sequences")
            reporter.complete("report", "report")
            return AnalysisBundle(mode=request.mode, input_type=request.input_type, inputs=[request.text]), {
                "stem": request.project_name
            }

        first_id = manager.submit(self._job_request("first"), runner)
        second_id = manager.submit(self._job_request("second"), runner)
        deadline = time.monotonic() + 2
        while time.monotonic() < deadline:
            by_id = {item.job_id: item for item in manager.snapshots()}
            if (
                by_id[first_id].status == "running"
                and by_id[first_id].progress > 0
                and by_id[second_id].status == "queued"
            ):
                break
            time.sleep(0.01)
        self.assertEqual(by_id[first_id].status, "running")
        self.assertEqual(by_id[second_id].status, "queued")
        self.assertGreater(by_id[first_id].progress, 0)
        self.assertEqual(
            [item.key for item in by_id[first_id].progress_items],
            ["mapping", "sequences", "lab_omics", "report"],
        )
        self.assertTrue(manager.cancel(first_id))
        gate.set()
        deadline = time.monotonic() + 3
        while time.monotonic() < deadline:
            by_id = {item.job_id: item for item in manager.snapshots()}
            if by_id[first_id].status == "cancelled" and by_id[second_id].status == "completed":
                break
            time.sleep(0.01)
        self.assertEqual(by_id[first_id].status, "cancelled")
        self.assertEqual(by_id[second_id].status, "completed")
        second_bundle, second_artifacts, _ = manager.get_result(second_id)
        self.assertEqual(second_bundle.inputs, ["Os01g0100100"])
        self.assertEqual(second_artifacts["stem"], "second")
        manager.shutdown()

    def test_background_progress_breakdown_tracks_children_and_warnings(self) -> None:
        manager = AnalysisJobManager(max_history=5)
        request = RiceGeneAnalysisRequest(
            project_name="progress detail",
            mode="单基因深度分析",
            input_type="RAP/MSU ID",
            text="Os01g0100100",
            selected_types=("CDS",),
            promoter_length=2000,
            transcript_scope=TRANSCRIPT_SCOPE_ALL,
            selected_predictors=("SignalP 6.0", "NLStradamus 1.8"),
            signalp_mode="fast",
            cnls_cutoff=5.0,
            nlstradamus_model=1,
            nlstradamus_cutoff=0.6,
            max_workers=1,
            include_ricedata=True,
            include_efp=True,
            efp_data_sources=("rice_rma", "ricestress_rma"),
        )

        def runner(job_request, reporter):
            reporter.complete("mapping", "mapped")
            reporter.complete("sequences", "sequences")
            reporter.complete("ricedata", "ricedata")
            reporter.update("efp", 1, 2, "first source")
            reporter.update_item("efp", "rice_rma", 1, 1, "matched")
            reporter.update("efp", 2, 2, "second source")
            reporter.update_item(
                "efp", "ricestress_rma", 1, 1, "request failed", warning=True
            )
            reporter.complete("efp", "charts", warning=True)
            reporter.complete("lab_omics", "lab omics")
            reporter.update("predictions", 1, 2, "SignalP")
            reporter.update_item("predictions", "SignalP 6.0", 1, 1, "matched")
            reporter.update("predictions", 2, 2, "NLStradamus")
            reporter.update_item(
                "predictions", "NLStradamus 1.8", 1, 1, "failed", warning=True
            )
            reporter.complete("predictions", "predictions", warning=True)
            reporter.complete("report", "report")
            return AnalysisBundle(
                mode=job_request.mode,
                input_type=job_request.input_type,
                inputs=[job_request.text],
                warnings=["mock service warning"],
            ), {}

        job_id = manager.submit(request, runner)
        deadline = time.monotonic() + 3
        while time.monotonic() < deadline:
            snapshot = next(item for item in manager.snapshots() if item.job_id == job_id)
            if snapshot.status in {"completed", "completed_with_warnings"}:
                break
            time.sleep(0.01)
        self.assertEqual(snapshot.status, "completed_with_warnings")
        self.assertEqual(
            [item.key for item in snapshot.progress_items],
            ["mapping", "sequences", "ricedata", "efp", "lab_omics", "predictions", "report"],
        )
        efp = next(item for item in snapshot.progress_items if item.key == "efp")
        predictions = next(item for item in snapshot.progress_items if item.key == "predictions")
        self.assertEqual(efp.status, "completed_with_warnings")
        self.assertEqual([child.status for child in efp.children], ["completed", "completed_with_warnings"])
        self.assertEqual(predictions.status, "completed_with_warnings")
        self.assertEqual(
            [child.status for child in predictions.children],
            ["completed", "completed_with_warnings"],
        )
        self.assertTrue(all(item.progress == 1.0 for item in snapshot.progress_items))
        manager.shutdown()

    def test_background_failure_marks_active_phase_and_retry_starts_clean(self) -> None:
        manager = AnalysisJobManager(max_history=5)
        attempts = {"count": 0}
        retry_gate = threading.Event()
        request = RiceGeneAnalysisRequest(
            project_name="phase failure",
            mode="单基因深度分析",
            input_type="RAP/MSU ID",
            text="Os01g0100100",
            selected_types=("CDS",),
            promoter_length=2000,
            transcript_scope=TRANSCRIPT_SCOPE_ALL,
            selected_predictors=(),
            signalp_mode="fast",
            cnls_cutoff=5.0,
            nlstradamus_model=1,
            nlstradamus_cutoff=0.6,
            max_workers=1,
            include_ricedata=False,
            include_efp=True,
            efp_data_sources=("rice_rma",),
        )

        def runner(job_request, reporter):
            attempts["count"] += 1
            reporter.complete("mapping", "mapped")
            reporter.complete("sequences", "sequences")
            if attempts["count"] == 1:
                reporter.update("efp", 0.95, 1, "plotting")
                reporter.complete_item("efp", "rice_rma", "fetched")
                raise AttributeError("mock plotting failure")
            retry_gate.wait(2)
            reporter.complete_item("efp", "rice_rma", "fetched")
            reporter.complete("efp", "charts")
            reporter.complete("report", "report")
            return AnalysisBundle(
                mode=job_request.mode,
                input_type=job_request.input_type,
                inputs=[job_request.text],
            ), {}

        failed_id = manager.submit(request, runner)
        deadline = time.monotonic() + 3
        while time.monotonic() < deadline:
            failed = next(item for item in manager.snapshots() if item.job_id == failed_id)
            if failed.status == "failed":
                break
            time.sleep(0.01)
        failed_efp = next(item for item in failed.progress_items if item.key == "efp")
        self.assertEqual(failed.stage, "efp")
        self.assertEqual(failed_efp.status, "failed")
        self.assertEqual(failed_efp.progress, 0.95)

        retry_id = manager.retry(failed_id)
        self.assertIsNotNone(retry_id)
        deadline = time.monotonic() + 3
        while time.monotonic() < deadline:
            retry = next(item for item in manager.snapshots() if item.job_id == retry_id)
            if retry.status == "running" and retry.stage == "sequences":
                break
            if retry.status == "running" and retry.stage == "efp":
                break
            time.sleep(0.01)
        retry_efp = next(item for item in retry.progress_items if item.key == "efp")
        self.assertEqual(retry_efp.progress, 0.0)
        self.assertEqual(retry_efp.status, "pending")
        retry_gate.set()
        deadline = time.monotonic() + 3
        while time.monotonic() < deadline:
            retry = next(item for item in manager.snapshots() if item.job_id == retry_id)
            if retry.status == "completed":
                break
            time.sleep(0.01)
        self.assertEqual(retry.status, "completed")
        manager.shutdown()

    def test_background_job_retry_after_failure(self) -> None:
        manager = AnalysisJobManager(max_history=5)
        attempts = {"count": 0}

        def runner(request, reporter):
            attempts["count"] += 1
            if attempts["count"] == 1:
                raise RuntimeError("mock failure")
            reporter.complete("mapping")
            reporter.complete("sequences")
            reporter.complete("report")
            return AnalysisBundle(mode=request.mode, input_type=request.input_type, inputs=[request.text]), {}

        failed_id = manager.submit(self._job_request("retryable"), runner)
        deadline = time.monotonic() + 2
        while time.monotonic() < deadline and manager.snapshots()[0].status != "failed":
            time.sleep(0.01)
        self.assertEqual(manager.snapshots()[0].status, "failed")
        retry_id = manager.retry(failed_id)
        self.assertIsNotNone(retry_id)
        deadline = time.monotonic() + 2
        while time.monotonic() < deadline:
            snapshot = next(item for item in manager.snapshots() if item.job_id == retry_id)
            if snapshot.status == "completed":
                break
            time.sleep(0.01)
        self.assertEqual(snapshot.status, "completed")
        self.assertEqual(attempts["count"], 2)
        manager.shutdown()


if __name__ == "__main__":
    os.chdir(APP_SOURCE)
    unittest.main(verbosity=2)
