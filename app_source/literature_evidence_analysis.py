"""Europe PMC metadata retrieval and source-separated genetic evidence import."""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup


EUROPE_PMC_URL = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
RAPDB_URL = "https://rapdb.dna.affrc.go.jp/viewer/gbrowse_details/irgsp1"
EVIDENCE_TAGS = {
    "knockout": r"\b(knockout|knock-out|CRISPR|loss.of.function)\b",
    "mutation": r"\b(mutant|mutation|allele)\b",
    "overexpression": r"\b(overexpression|over-expression|transgenic)\b",
    "RNAi": r"\b(RNAi|RNA interference|silencing)\b",
    "QTL": r"\bQTL\b", "GWAS": r"\bGWAS\b",
    "interaction": r"\b(interact|binding|two-hybrid|pull-down)\b",
    "expression": r"\b(expression|transcript|RNA-seq)\b",
}


def _tags(text: str) -> str:
    return ",".join(label for label, pattern in EVIDENCE_TAGS.items() if re.search(pattern, text, re.I))


def build_query(identifiers: list[str]) -> str:
    terms = [f'"{value}"' for value in dict.fromkeys(value.strip() for value in identifiers if value.strip())]
    return f"({' OR '.join(terms)}) AND (rice OR \"Oryza sativa\")" if terms else ""


def fetch_europe_pmc(targets: list[dict[str, object]], max_results: int = 50, session: requests.Session | None = None) -> tuple[list[dict[str, object]], dict[str, bytes], list[str]]:
    client = session or requests.Session(); rows, raw, warnings = [], {}, []
    queried_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    for target in targets:
        aliases = [str(target.get(key) or "") for key in ("rap_gene", "msu_id", "gene_symbol", "gene_name")]
        aliases.extend(target.get("aliases") or [])
        query = build_query(aliases)
        if not query:
            continue
        try:
            response = client.get(EUROPE_PMC_URL, params={"query": query, "format": "json", "pageSize": max_results, "resultType": "core"}, timeout=45)
            response.raise_for_status(); payload = response.json()
            key = re.sub(r"[^A-Za-z0-9._-]+", "_", str(target.get("rap_gene") or target.get("input_id") or "query"))
            raw[f"{key}/europe_pmc.json"] = response.content
            for item in payload.get("resultList", {}).get("result", []):
                title = str(item.get("title") or ""); abstract = str(item.get("abstractText") or "")
                context = f"{title} {abstract}"
                matched = [alias for alias in aliases if alias and re.search(rf"(?<![A-Za-z0-9_]){re.escape(alias)}(?![A-Za-z0-9_])", context, re.I)]
                if not matched:
                    continue
                rows.append({
                    "input_id": target.get("input_id", ""), "rap_gene": target.get("rap_gene", ""), "msu_id": target.get("msu_id", ""),
                    "pmid": item.get("pmid", ""), "doi": item.get("doi", ""), "title": title,
                    "year": item.get("pubYear", ""), "journal": item.get("journalTitle", ""),
                    "authors": item.get("authorString", ""), "abstract_available": bool(abstract),
                    "abstract_text": abstract,
                    "matched_fields": ",".join(matched), "evidence_tags": _tags(context),
                    "verification_status": "需人工核对全文", "source_type": "online_metadata",
                    "source_url": f"https://europepmc.org/article/MED/{item.get('pmid')}" if item.get("pmid") else response.url,
                    "queried_at": queried_at, "parameters": query, "status": "matched", "error": "",
                })
        except Exception as exc:
            warnings.append(f"Europe PMC 查询失败（{target.get('input_id')}）：{type(exc).__name__}: {exc}")
    deduped, seen = [], set()
    for row in rows:
        key = (str(row.get("pmid") or "").lower(), str(row.get("doi") or "").lower())
        key = key if any(key) else (str(row.get("title") or "").casefold(), "")
        if key not in seen: seen.add(key); deduped.append(row)
    return deduped, raw, warnings


def enrich_ricedata_references(
    references: list[dict[str, object]],
    session: requests.Session | None = None,
) -> tuple[list[dict[str, object]], dict[str, bytes], list[str]]:
    """Complete PMID and stable metadata for RiceData-linked DOI records."""
    client = session or requests.Session()
    enriched: list[dict[str, object]] = []
    raw: dict[str, bytes] = {}
    warnings: list[str] = []
    queried_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    for reference in references:
        row = dict(reference)
        doi = str(row.get("doi") or "").strip()
        row.setdefault("verification_status", "RiceData 关联文献，需核验与当前证据的具体关系")
        row.setdefault("matched_by", "ricedata_reference_id")
        row["queried_at"] = queried_at
        if not doi:
            enriched.append(row)
            continue
        try:
            response = client.get(
                EUROPE_PMC_URL,
                params={"query": f'DOI:"{doi}"', "format": "json", "pageSize": 5, "resultType": "core"},
                timeout=45,
            )
            response.raise_for_status()
            key = re.sub(r"[^A-Za-z0-9._-]+", "_", str(row.get("reference_id") or doi))
            raw[f"ricedata_reference_{key}/europe_pmc.json"] = response.content
            results = response.json().get("resultList", {}).get("result", [])
            matched = next(
                (item for item in results if str(item.get("doi") or "").casefold() == doi.casefold()),
                results[0] if results else None,
            )
            if matched:
                row["pmid"] = row.get("pmid") or matched.get("pmid", "")
                row["title"] = row.get("title") or matched.get("title", "")
                row["year"] = row.get("year") or matched.get("pubYear", "")
                row["journal"] = matched.get("journalTitle", "")
                row["authors"] = matched.get("authorString", "")
                row["abstract_available"] = bool(matched.get("abstractText"))
                row["abstract_text"] = str(matched.get("abstractText") or "")
                row["europe_pmc_url"] = (
                    f"https://europepmc.org/article/MED/{matched.get('pmid')}"
                    if matched.get("pmid") else response.url
                )
                row["status"] = "matched"
            else:
                row["status"] = row.get("status") or "not_found"
        except Exception as exc:
            warnings.append(f"RiceData reference {row.get('reference_id') or doi} 的 PMID 补齐失败：{type(exc).__name__}: {exc}")
            row["error"] = f"{type(exc).__name__}: {exc}"
        enriched.append(row)
    return enriched, raw, warnings


def import_manual_evidence(payload: bytes, filename: str) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    if not payload:
        return records
    if Path(filename).suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        values = list(sheet.iter_rows(values_only=True))
        if values:
            headers = [str(value or f"column_{index}") for index, value in enumerate(values[0], 1)]
            records = [dict(zip(headers, row)) for row in values[1:] if any(value not in (None, "") for value in row)]
    else:
        text = payload.decode("utf-8-sig", errors="replace")
        records = list(csv.DictReader(io.StringIO(text)))
    queried_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    return [{**row, "source_type": "manual_import", "source_url": f"uploaded://{filename}", "queried_at": queried_at, "verification_status": str(row.get("verification_status") or "人工证据"), "status": str(row.get("status") or "imported"), "error": ""} for row in records]


def genetic_evidence_from_ricedata(
    rows: list[dict[str, object]],
    references: list[dict[str, object]] | None = None,
) -> list[dict[str, object]]:
    output = []
    references_by_gene: dict[str, list[dict[str, object]]] = {}
    for reference in references or []:
        gene_id = str(reference.get("gene_id") or "")
        references_by_gene.setdefault(gene_id, []).append(reference)
    for row in rows:
        for key, tag in (("突变体表型", "mutation"), ("定位与克隆", "QTL/cloning"), ("生物学功能", "functional annotation")):
            if row.get(key):
                evidence_text = str(row[key])
                candidate_references = references_by_gene.get(str(row.get("GeneID") or ""), [])
                linked: list[dict[str, object]] = []
                for reference in candidate_references:
                    year = str(reference.get("year") or "")
                    authors = str(reference.get("authors") or "")
                    first_author = authors.split(",", 1)[0].split(" ", 1)[0] if authors else ""
                    matched_by = []
                    if year and year in evidence_text:
                        matched_by.append("citation_year")
                    if first_author and re.search(rf"\b{re.escape(first_author)}\b", evidence_text, re.I):
                        matched_by.append("citation_author")
                    linked.append({**reference, "evidence_match": "+".join(matched_by) if matched_by else "ricedata_gene_page"})
                direct = [item for item in linked if "citation_year" in str(item.get("evidence_match"))]
                linked_dois = [str(item.get("doi") or "") for item in (direct or linked) if item.get("doi")]
                if direct:
                    verification = "直接支持（RiceData 证据文本与关联引用年份匹配）"
                    matched_by = ",".join(sorted({str(item.get("evidence_match")) for item in direct}))
                elif linked:
                    verification = "RiceData 关联文献，需核验与该证据的具体关系"
                    matched_by = "ricedata_gene_page"
                else:
                    verification = "数据库已知证据，未解析到关联文献"
                    matched_by = "ricedata_record"
                output.append({
                    "input_id": row.get("check", ""),
                    "rap_gene": row.get("RAP_Locus", ""),
                    "msu_id": row.get("MSU_Locus", ""),
                    "gene_symbol": row.get("GeneSymbol", ""),
                    "evidence_type": tag,
                    "evidence_text": evidence_text,
                    "linked_dois": ",".join(dict.fromkeys(linked_dois)),
                    "linked_reference_ids": ",".join(str(item.get("reference_id") or "") for item in (direct or linked)),
                    "matched_by": matched_by,
                    "verification_status": verification,
                    "source_type": "RiceData",
                    "source_url": row.get("source_url", ""),
                    "status": "database_record",
                    "error": "",
                })
    return output


def fetch_rapdb_genetic_evidence(rap_ids: list[str], session: requests.Session | None = None) -> tuple[list[dict[str, object]], dict[str, bytes], list[str]]:
    """Preserve RAP-DB allele/mutant text as database evidence; never infer a mechanism."""
    client = session or requests.Session(); evidence, raw, warnings = [], {}, []
    queried_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    keywords = re.compile(r"allele|variation|variant|mutation|mutant|phenotype|QTL", re.I)
    for rap_id in dict.fromkeys(value for value in rap_ids if value):
        try:
            response = client.get(RAPDB_URL, params={"name": rap_id}, timeout=(8, 40), headers={"User-Agent": "MyBioTools/1.9.7"})
            response.raise_for_status(); raw[f"{rap_id}/rapdb_gene_detail.html"] = response.content
            soup = BeautifulSoup(response.text, "html.parser"); matched = 0
            for row in soup.find_all("tr"):
                cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["th", "td"])]
                text = " | ".join(value for value in cells if value)
                if text and keywords.search(text):
                    evidence.append({"input_id": rap_id, "rap_gene": rap_id, "msu_id": "", "evidence_type": "RAP-DB allele/mutation record", "evidence_text": text, "verification_status": "数据库已知证据，需核对原始引用", "source_type": "RAP-DB", "source_url": response.url, "queried_at": queried_at, "status": "database_record", "error": ""}); matched += 1
            if not matched:
                warnings.append(f"{rap_id} RAP-DB 未取得可识别的 allele/mutation 记录。")
        except Exception as exc:
            warnings.append(f"{rap_id} RAP-DB 遗传证据查询失败：{type(exc).__name__}: {exc}")
    return evidence, raw, warnings
